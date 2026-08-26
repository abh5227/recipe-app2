#!/usr/bin/env python3
"""build_library.py: build the first-pass ingredient library from join.db and vocab/.

    python3.13 build_library.py            -> previews/ingredient-list-pass1.xlsx

WHAT THIS PRODUCES. One row per ingredient: a canonical name, every variation that
resolves to it with the language and source and field each came from, what kind of thing
it is, and a confidence with the reason spelled out. NAMES AND VARIATIONS ONLY. No
descriptions, no facts, no densities, no substitutions. Those come later and writing them
before the list is right would be writing them twice.

⚠️ NO NAME IS INVENTED HERE. Every canonical name and every variation is copied verbatim
from a source. The only things this file decides are WHICH entries exist and WHICH name
leads. Both decisions are stated below and both are judgements.

WHAT IT READS
    join.db            the name buckets, built by build_join.py, git-ignored
    sources.db         read-only, REQUIRED. Entry names and English glosses. git-ignored.
                       ⚠️ The build STOPS without it rather than returning a shorter list.
                       Measured: an earlier draft treated it as optional and silently
                       returned 11,769 entries instead of 11,153, because the E-number
                       filter reads OFF's stored names and matched nothing.
    vocab/*.json       the classification model. ⚠️ CANNOT BE REGENERATED, see vocab/README.md
    ingredient_cuts.py the cut rules and the override list
    reviewed.py        the 330 hand-read verdicts
    hand_removals.csv  ⚠️ Andy's removals. THE DECISION LIVES THERE, NOT IN THE SHEET,
                       because the sheet is regenerated. Marked in the spreadsheet and
                       pulled back by harvest_marks.py.

WHAT IS REPRODUCIBLE AND WHAT IS NOT. Given the same join.db and the same vocab/ this
script is deterministic and rebuilds the sheet exactly. What it CANNOT rebuild is the
judgement in ingredient_cuts.py and reviewed.py, which is why both are committed as data
rather than recomputed here.

THE FOUR ADMISSION RULES. Only Wikidata and Open Food Facts may create an entry,
because only they carry anything resembling a food classification.

  1  Wikidata classifies the item "Ingredient or foodstuff".                6,605
  2  Wikidata carries NO classification, but an OFF ingredients-taxonomy       487
     entry shares its name.
     ⚠️ THE WEAKEST RULE, and it exists for one measured reason: the Wikidata item that
     gets gochugaru right carries no food category at all. It also admitted dish, diet,
     seed, legume, flavoring and biscuit, which is why the sheet marks all 487.
  3  An OFF ingredients-taxonomy entry reaching no Wikidata item, after the    4,269
     json and txt copies of one concept are merged.
  4  THE DRINKS RULE. Wikidata calls the item a Drink and NOT an ingredient,           22
     and an OFF ingredients-taxonomy entry's ENGLISH canonical_name EQUALS the
     item's ENGLISH label. See drinks_rule for the measurement.
     ⚠️ NUMBERED LAST, MEASURED SECOND. It is wrong on none of its 22 where rule 2
     admitted dish, diet and biscuit. The numbers are quoted across ingredient_cuts.py,
     reviewed.py and the sheet, so rule 2 keeps its number rather than being demoted.

⚠️ AGROVOC, WIKTIONARY AND WIKIPEDIA NEVER CREATE AN ENTRY. They supply variations only.
None carries a food classification: AGROVOC has no type field in the store and its entries
begin 2,4-D and A horizons, and the two Wikimedia sources are a dictionary and an
article-redirect list. Measured on 60 hand-read Wiktionary candidates, its topics-and-
categories signal is 57 to 75% wrong AND misses Shaoxing wine, so there is nothing better
to use than the five-entry override list.

OWNERSHIP, AND WHY IT IS NOT "EVERY NAME IN MY BUCKETS". Every source entry belongs to at
most ONE library entry. Seeds own themselves, and every other source entry goes to the
anchor it shares the most buckets with.

    Without this the ginger anchor swallowed all six rows of the 'hing' bucket. That
    bucket holds two different things: Wikidata, OFF and its canonical_name all label
    ginger "hing" in ZHUANG, while AGROVOC's altLabel and OFF 4698's synonym are ENGLISH
    and mean asafoetida. Ownership gives each entry the rows its own sources wrote, and
    the language column on every variation is what makes the split readable.
"""
import collections, json, os, pickle, re, sqlite3, sys, unicodedata

import csv

import ingredient_cuts as CUTS
# ⚠️ THE SAME KEY THE JOIN USES, AND THE SAME KEY A RECIPE LABEL WILL BE MATCHED ON.
#    Two rows answering to one name is only a defect under the key the lookup will use,
#    so the name rules below key on norm_name rather than on casefold(). Measured:
#    casefold finds 1,260 pairs where norm_name finds 1,301, a difference of 41 that is
#    entirely hyphens, brackets and apostrophes.
from build_join import norm_name
try:
    import reviewed
except ImportError:                                   # the verdicts are optional to run
    reviewed = None

HAND_REMOVALS = os.environ.get("HAND_REMOVALS", "hand_removals.csv")
AUTHORED_ROWS = os.environ.get("AUTHORED_ROWS", "authored_rows.csv")
HAND_RENAMES = os.environ.get("HAND_RENAMES", "hand_renames.csv")
JOIN_DB = os.environ.get("JOIN_DB", "join.db")
SOURCES_DB = os.environ.get("SOURCES_DB", "sources.db")
VOCAB = os.environ.get("VOCAB_DIR", "vocab")
OUT = os.environ.get("LIBRARY_XLSX", "previews/ingredient-list-pass1.xlsx")

SOURCE_NAME = {"wikidata": "Wikidata", "off_taxonomy": "OFF", "agrovoc": "AGROVOC",
               "wiktextract": "Wiktionary", "wikipedia_redirect": "Wikipedia"}
# The source's own display name for an entry, per source's own vocabulary.
PRIMARY_KINDS = {"word", "label", "prefLabel", "canonical_name", "name", "article_title"}
# ⚠️ The fields the measured errors came from. 20.8% wrong on the 200-bucket sample.
ALIAS_KINDS = {"synonym", "alt_of", "form", "altLabel"}
FIELD_NAME = {"synonym": "synonym", "alt_of": "alternative form", "form": "word form",
              "altLabel": "alias", "alias": "alias", "redirect": "redirect",
              # ⚠️ NO SOURCE WROTE THIS ONE. See build_library.strip_as_food.
              "derived": "DERIVED HERE, no source wrote it",
              **{k: "primary name" for k in PRIMARY_KINDS}}

INGREDIENT = "Ingredient or foodstuff"
# ⚠️ Excluded even when Wikidata ALSO calls the item an ingredient. 179 brands and 34
#    fictional items were admitted by kind and removed anyway. Dish, Taxon, Chemical and
#    Drink are NOT here, because ingredient-plus-those is the dual nature and is kept.
FATAL_KINDS = {"Fictional food": "fictional",
               "Brand or trademark": "a brand or trademark"}
# ⚠️ MARKED, NEVER MOVED, AND THE MEASUREMENT IS IN ingredient_cuts.DECLINED.
#    653 rows carry one of these kinds AND "Ingredient or foodstuff". Reading 40 found
#    16 dishes and 24 ingredients, so a batch move would take chocolate, shrimp, honey,
#    ham, soy sauce and 387 others out of the library. The flag makes the 653 sortable
#    so the separation happens by reading, one row at a time.
DISH_KINDS = {"Dish or prepared food", "Cuisine, recipe or meal"}
# ⚠️ RULE 4 ADMITS FROM THE FIRST AND NEVER FROM THE SECOND. See drinks_rule.
DRINK = "Drink"
APPELLATION = "Appellation or growing region"
# ⚠️ ONE FIELD PER OFF COPY, NOT TWO OPINIONS. See english_primary.
OFF_PRIMARY = {"canonical_name", "name"}
# ⚠️ WORDS THAT NAME A CONCENTRATION RATHER THAN A DIFFERENT INGREDIENT.
#    See mark_strength. Deliberately not a complete list of preparation words: 'smoked'
#    and 'pickled' make a different ingredient, 'powder' and 'paste' make the same one at
#    a different strength, and only the second kind belongs here.
STRENGTH_WORDS = {"paste", "concentrate", "extract", "essence", "juice", "water", "powder",
                  "puree", "purée", "syrup", "granules", "cube", "dried", "fresh", "raw",
                  "ground", "minced", "liquid", "condensed", "evaporated", "flakes",
                  "desiccated", "milk", "cream", "flour"}
# ⚠️ AN ENTRY WHOSE WORD IS AN INITIALISM. 'en:MPFE:noun#0' -> 'MPFE'. See
#    drop_initialism_expansions, which is the only thing that reads this.
INITIALISM = re.compile(r"^[A-Z]{2,6}$")
E_NUMBER = re.compile(r"^e\s?\d{3}", re.I)
PROCESS_NAMES = {"frying", "cooking", "baking", "smoking", "drying", "fermentation",
                 "roasting", "preparation"}
BINOMIAL = re.compile(r"^[A-Z][a-z]+ [a-z]+(?: (?:subsp|var|f)\.? [a-z]+)?$")


def is_binomial(canonical, variations):
    """⚠️ SHAPE IS NOT EVIDENCE, AND ON ITS OWN IT WAS WRONG ABOUT 9 NAMES IN 10.

    BINOMIAL alone is a capital word followed by a lowercase one, which is the shape of
    every English name carrying a proper noun in front of it. It fired on 705 rows.
    Reading 60 of them found 5 genuine binomials, 8.3%, 95% CI [3.6%, 18.1%], so roughly
    646 of the 705 were wrong. Kaiser roll, Swiss roll, Welsh onion, Atlantic salmon,
    Shaoxing wine, Napa cabbage, Cheddar cheese, Serrano ham and Manuka honey all fired.

    ⚠️ THE FLAG IS NOT DECORATION. It tells the reader to pick a cook's name out of the
    variations by hand, so a wrong flag is work that should never have been started.

    The evidence the shape lacked is a source calling the string Latin. 56 of the 705
    carry a Latin language tag on the canonical itself, and the hand-read 5 were all 5 of
    them. Known cost, both directions:

      one false positive   'American cheese'. la.wikipedia did not translate the name, so
                           Wikidata carries it as a Latin label. Real tag, real name,
                           not a binomial.
      about fifteen misses mostly bacterial culture names that no source tags as Latin.
                           Lactobacillus acidophilus, Bifidobacterium longum,
                           Streptococcus thermophilus, Tuber magnatum, Tamarindus indica.
                           ⚠️ A MISS COSTS LESS THAN A FALSE POSITIVE HERE. Nobody writes
                           Lactobacillus acidophilus in a recipe, so an unflagged one
                           creates no work. A wrong flag does.

    ⚠️ A GENUS VOCABULARY WAS TRIED AND MADE IT WORSE, recorded so it is not retried.
    Deriving genus names from the corpus's own Latin-tagged strings yielded 339 genera and
    added 14 rows, only 5 of them real. 'American cheese' put 'American' in the genus list
    and pulled in American lobster, American mustard, American beef and three more. One
    bad tag propagated. Requiring the genus twice and the epithet absent from an English
    wordlist cut it to 4 correct additions, at the price of a system wordlist this repo
    does not otherwise need."""
    if not BINOMIAL.match(canonical):
        return False
    return any(lang.startswith("la") for _, _, lang in variations.get(canonical, ()))


def load_vocab():
    """The classification model. ⚠️ These files cannot be regenerated, see vocab/README.md."""
    def read(name, key):
        with open(os.path.join(VOCAB, name), encoding="utf-8") as fh:
            return json.load(fh)[key]
    return (read("wikidata-kinds.json", "kinds"),
            read("wikidata-superclasses.json", "superclasses"),
            read("off-taxonomy-tree.json", "parents"))


def off_tree(src):
    """Rebuild the OFF parent tree from sources.db when it is present. Falls back to the
    committed copy, which exists because sources.db is 5.18 GB and git-ignored."""
    if src is None:
        return None
    parents = collections.defaultdict(set)
    rows = src.execute("SELECT e.entry_id, e.raw, f.dataset FROM off_taxonomy_entry e "
                       "JOIN source_fetch f ON f.id = e.fetch_id")
    for entry_id, raw, dataset in rows:
        if dataset.endswith("_json"):
            try:
                got = json.loads(raw).get("parents") or []
            except ValueError:
                got = []
        else:
            got = re.findall(r"<\s*(en:[\w\-']+)", raw or "")
        for parent in got:
            parents[entry_id].add(str(parent))
    return {k: sorted(v) for k, v in parents.items()}


def read_members(join):
    """Every member row, grouped by source entry and by bucket."""
    by_entry = collections.defaultdict(list)
    by_bucket = collections.defaultdict(list)
    for norm, source, dataset, entry_id, kind, lang, text, _ in join.execute(
            "SELECT norm, source, dataset, entry_id, kind, lang, text, via_override "
            "FROM join_member"):
        by_entry[(source, dataset, entry_id)].append((norm, kind, lang, text))
        by_bucket[norm].append((source, dataset, entry_id, kind, lang, text))
    return by_entry, by_bucket


def english_primary(by_entry):
    """The English primary name each source entry states, read off the join's own members.

    ⚠️ NOT A SECOND QUERY AGAINST sources.db, AND THE TWO WERE CHECKED EQUAL BEFORE THIS
    WAS WRITTEN. The join carries the Wikidata label and the OFF English name as ordinary
    member rows, so both derivations admit exactly the same 22 drinks. Reading the join
    keeps rule 4 inside the same data every other rule is decided on.

    ⚠️ BOTH OFF FIELDS, BECAUSE OFF IS LOADED TWICE AND THE TWO COPIES DISAGREE ON THE
    FIELD NAME. The txt taxonomy writes canonical_name and the json taxonomy writes name.
    Measured: 4,699 txt entries carry an English canonical_name and ZERO of the 6,442
    json entries do, so reading canonical_name alone sees one copy of every concept and
    misses the other. It does not change who is admitted, and that was checked rather
    than assumed: canonical_name alone and canonical_name-or-name admit the same 22. It
    changes what those 22 ABSORB, which is what left a duplicate 'black tea' row behind.
    """
    wikidata, off = {}, collections.defaultdict(set)
    for (source, dataset, entry_id), rows in by_entry.items():
        for _, kind, lang, text in rows:
            if (lang or "").lower() != "en":
                continue
            if source == "wikidata" and kind == "label":
                wikidata.setdefault(entry_id, text)
            elif source == "off_taxonomy" and kind in OFF_PRIMARY:
                off[norm_name(text)].add((dataset, entry_id))
    return wikidata, off


def drinks_rule(by_entry, kinds, admitted):
    """RULE 4. A Wikidata Drink carrying no ingredient kind is admitted when an Open Food
    Facts ingredient entry's ENGLISH canonical_name EQUALS the item's ENGLISH label.

    ⚠️ THE NARROWING IS THE RULE, AND THE OBVIOUS VERSION WAS MEASURED AS A COIN FLIP.
    Admitting every excluded drink that merely SHARES A NAME BUCKET with an OFF ingredient
    entry admits 65 and is right on 39, which is 60%. Requiring the two English names to
    be EQUAL admits 22 and is wrong on none, 91% strict and 100% counting cider and
    sparkling wine as ingredients. All 65 are read one at a time in reviewed.DRINKS_SAMPLE.

    ⚠️ DO NOT RELAX THIS BACK TO SHARED BUCKETS. What the loose version lets in is a
    cross-language homograph every time, and that is the THIRD time the same failure has
    landed in this pipeline: 'ni' is nickel and it is milk, 'gula' is sugar and it is yolk,
    'granada' is a city and it is a pomegranate. The drinks repeat it exactly. 'latte'
    shares a bucket with OFF milk, 'Uva' with grape, 'Doogh' with dough, 'Posca' with
    vinegar, 'Turkish coffee' with flour, and 'weak coffee' with WATER on a bucket carrying
    33 recipe lines. Equal English names kill all six and cost nothing.

    ⚠️ AND DO NOT POINT IT AT APPELLATIONS. The same narrowing over the 1,034 excluded
    appellations admits six, four of which are wine regions, so the appellation kind is
    excluded here explicitly rather than by luck. It blocks exactly one item, Cava.

    ⚠️ WHAT IT DOES NOT ADMIT IS RECORDED, NOT LOST. 19 of the 65 read as real ingredients
    and fail the name test because their OFF match is a PARENT rather than the same thing:
    hot chocolate matches cocoa, horchata matches tigernut milk, kvass matches sourdough,
    Hibiscus tea matches roselle flower. Dropping masala chai to tea and drip coffee to
    coffee is correct. Those four are an outstanding reading, listed in
    reviewed.DRINKS_SAMPLE under 'outstanding', and they are readable at leisure.
    """
    wikidata, off = english_primary(by_entry)
    out = set()
    for source, _, entry_id in by_entry:
        if source != "wikidata" or entry_id in admitted:
            continue
        item = kinds.get(entry_id, {}).get("kinds", {})
        if DRINK not in item or INGREDIENT in item or APPELLATION in item:
            continue
        label = norm_name(wikidata.get(entry_id, "") or "")
        if label and label in off:
            out.add(entry_id)
    return out


def pick_anchors(by_entry, by_bucket, kinds):
    """The four admission rules. Returns (rule1, rule2, drinks, off_only_groups)."""
    wd_in_join = {e for (s, d, e) in by_entry if s == "wikidata"}
    off_in_join = {(d, e) for (s, d, e) in by_entry if s == "off_taxonomy"}

    wd_with_off = set()
    for rows in by_bucket.values():
        if any(r[0] == "off_taxonomy" for r in rows):
            wd_with_off |= {r[2] for r in rows if r[0] == "wikidata"}

    rule1 = {q for q in wd_in_join if INGREDIENT in kinds.get(q, {}).get("kinds", {})}
    rule2 = {q for q in wd_in_join if not kinds.get(q, {}).get("kinds")} & wd_with_off
    drinks = drinks_rule(by_entry, kinds, rule1 | rule2)

    covered = set()
    for q in rule1 | rule2:
        for norm, *_ in by_entry[("wikidata", "food_items_q2095", q)]:
            covered |= {(d, e) for (s, d, e, *_) in by_bucket[norm] if s == "off_taxonomy"}
    # ⚠️ A RULE-4 ANCHOR COVERS EXACTLY THE OFF ENTRY IT WAS ADMITTED ON, AND NOTHING
    #    ELSE. Covering suppresses the OFF-only row for an entry a Wikidata anchor already
    #    reaches, and it is computed over EVERY bucket the anchor touches. Measured on the
    #    22 drinks, that width is a defect in both directions:
    #      letting them cover everything they touch DELETED SIX EXISTING ROWS. Italian
    #      'melù' is blue whiting the FISH and also a wine, so the wine anchor swallowed
    #      the fish's row. Spanish 'Espumante' is sparkling wine and also OFF's additive
    #      'Foaming agent'. And 'champagne', the appellation kept out of rule 4 on
    #      purpose, arrived as a variation while its own row was deleted.
    #      letting them cover NOTHING left 12 names carried by two rows, one Wikidata and
    #      one OFF, which is the defect 'one name, one row' was written to remove.
    #    The same equal-English-names test that admits the anchor decides what it absorbs,
    #    so it takes en:black-tea and leaves the fish alone.
    wd_label, off_by_name = english_primary(by_entry)
    for q in drinks:
        covered |= off_by_name.get(norm_name(wd_label.get(q, "") or ""), set())

    # ⚠️ OFF IS LOADED TWICE, as a json taxonomy and a txt taxonomy, so most concepts
    #    appear as TWO entries. Two OFF entries sharing a bucket are one concept.
    rest, groups, seen = off_in_join - covered, [], set()
    for key in sorted(rest):
        if key in seen:
            continue
        group, stack = {key}, [key]
        while stack:
            dataset, entry_id = stack.pop()
            for norm, *_ in by_entry[("off_taxonomy", dataset, entry_id)]:
                for s2, d2, e2, *_ in by_bucket[norm]:
                    if s2 == "off_taxonomy" and (d2, e2) in rest and (d2, e2) not in group:
                        group.add((d2, e2))
                        stack.append((d2, e2))
        seen |= group
        groups.append(sorted(group))
    return rule1, rule2, drinks, groups


# ⚠️ ONE PRIMARY NAME PER CONCEPT PER LANGUAGE, AND THE SOURCES KEEP THAT PROMISE EXACTLY.
#    Measured, not assumed: 0 of 250,765 AGROVOC (entry, language) pairs hold two
#    prefLabels, 0 of 238,997 Wikidata pairs hold two labels, 0 of 61,561 Open Food Facts
#    pairs hold two canonical_names and 0 of 86,527 hold two names.
#
#    So two DIFFERENT primary names from one source in one language on one library row is
#    two source concepts merged, and no name-shape test is involved in saying so.
#
# ⚠️ BOTH OFF FIELDS, AND IT SHIPPED READING ONE. Open Food Facts is loaded twice and
#    the two copies disagree on the field name: the txt taxonomy writes canonical_name and
#    the json taxonomy writes name, 4,699 of 5,590 against 0, and 0 against 5,515 the
#    other way. Reading canonical_name alone saw one copy of every concept, so this rule
#    was half-blind from the commit that introduced it. See docs/measuring-the-premise.md,
#    case 7.
#
#    Measured cost of the half-view, which is small and is not nothing: 8 more rows carry
#    a second primary name (938 to 946) and 558 more marks appear, all of them on rows
#    carrying ZERO recipe lines. Six more names move, of which one carries a line, and
#    it is a wrong-traffic case: 'red bell pepper' leaves the bell pepper row for its own.
#    NO hand-read verdict in reviewed.py is overturned and nothing stops moving.
PRIMARY_CLAIM = {("agrovoc", "prefLabel"), ("wikidata", "label"),
                 ("off_taxonomy", "canonical_name"), ("off_taxonomy", "name")}


def primary_claims(rows, source):
    """The (source, language) slot -> the primary name an entry states in it.

    ⚠️ THE NAME, NOT JUST THE SLOT, AND THE FIRST DRAFT KEPT ONLY THE SLOT. It marked any
    second primary name in an occupied language, so two entries that AGREE got flagged as a
    merge: 'rosemary' was marked an intruder on the rosemary row. The premise is two
    DIFFERENT primary names. Measured, the correction removes 914 of 15,850 marks, 5.8%,
    which is smaller than it looks worth: the marks it removes are the ones that would have
    read as nonsense to anyone checking the column."""
    return {(source, (lang or "").lower()): text
            for _, kind, lang, text in rows if (source, kind) in PRIMARY_CLAIM}


def assign_ownership(entries, by_entry, by_bucket):
    """Every source entry belongs to at most ONE library entry. See the module docstring
    for the hing bucket, which is why this exists.

    ⚠️ AND AN ENTRY CARRYING A PRIMARY NAME DOES NOT JOIN AN ANCHOR THAT ALREADY HOLDS ONE
    FROM THE SAME SOURCE IN THE SAME LANGUAGE. Without this, 954 rows carried two or more
    concepts of one source over 698 recipe lines. egg yolk held sugar AND Amanita caesarea,
    milk held nickel, cabbage held water, cream held Panax, honey held common sole.

    The cause is that the "most shared buckets" tie-break has no floor and one homograph
    is enough to win. milk absorbed AGROVOC's nickel on the single bucket 'ni', the chemical
    symbol, out of 263 buckets. egg yolk absorbed Open Food Facts' sugar on the single
    bucket 'gula', which is sugar in Malay, out of 147.

    ⚠️ A SHARED-BUCKET FLOOR WAS MEASURED AND REJECTED. 17,933 of 23,814 absorbed entries,
    75.3%, share exactly ONE bucket with their owner, so a floor of two would un-own three
    quarters of the library's variation coverage to fix 954 rows. The precision is in the
    field rather than the count: of entries that carry a primary name, only 32.1% won on a
    single bucket. This rule has no threshold in it.

    ⚠️ NOTHING IS UN-OWNED AND NOTHING IS RESHUFFLED. An unowned entry is a name nothing can
    reach, which is worse than a name on the wrong row. What the rule marks is the intruding
    PRIMARY NAME, and resolve_borrowed then takes it off the row only where another row
    carries it, so no name can be lost. The entry's other names stay: Open Food Facts' sugar
    entry keeps contributing its 90 translations to whatever row it landed on, and only the
    word 'sugar' stops answering for egg yolk.

    ⚠️ RESHUFFLING WAS BUILT FIRST AND MEASURED AND IT DOES NOT WORK. Sending the entry to
    its next-best free anchor moved 954 merged rows to 931 and put recipe lines UP from 698
    to 721, because 1,727 of 1,914 displaced entries had no free anchor at all. A Wikidata
    anchor's own seed already holds a label in every language it is labeled in, so a second
    Wikidata item is blocked from every candidate and lands back where it started. The 187
    that did move took their clash with them to a new row.

    Entries are placed strongest-claim-first, so the ordering is deterministic rather than
    dictionary order, in which an entry sharing one bucket could take the slot from one
    sharing forty."""
    # ⚠️ THE ARTICLE A REDIRECT ENTRY IS. A wikipedia_redirect entry is ONE en.wikipedia
    #    article plus every name that redirects to it, and build_join already stores the
    #    article under its own kind on the same entry_id. 12,055 entries, all 12,055 carrying
    #    an article_title, against 42,762 redirect names. Nothing here is fetched or joined
    #    again: the signal has been sitting in join.db unread.
    article = {}
    for key, rows in by_entry.items():
        if key[0] != "wikipedia_redirect":
            continue
        for _, kind, _, text in rows:
            if kind == "article_title":
                article[key] = text
                break

    owner, seeds, held = {}, {}, collections.defaultdict(dict)
    for i, entry in enumerate(entries):
        for key in entry["seed"]:
            owner[key] = seeds[key] = i
            held[i].update(primary_claims(by_entry[key], key[0]))
    shared = collections.defaultdict(collections.Counter)
    for i, entry in enumerate(entries):
        for bucket in entry["buckets"]:
            for source, dataset, entry_id, *_ in by_bucket[bucket]:
                if (source, dataset, entry_id) not in seeds:
                    shared[(source, dataset, entry_id)][i] += 1

    # ⚠️ STRONGEST CLAIM FIRST, AND THE SORT IS THE RULE'S DETERMINISM. Placing entries in
    #    dict order would let an entry sharing one bucket take the slot from one sharing forty.
    order = sorted(shared, key=lambda k: (-max(shared[k].values()), k))
    intruders = collections.defaultdict(set)
    for key in order:
        best = max(shared[key].items(), key=lambda kv: (kv[1], -kv[0]))[0]
        owner[key] = best
        wants = primary_claims(by_entry[key], key[0])
        if not wants:
            continue                                  # variations only, nothing to clash
        for slot, text in wants.items():
            sitting = held[best].get(slot)
            if sitting is not None and norm_name(sitting) != norm_name(text):
                intruders[best].add(text)             # a SECOND, DIFFERENT primary name
            held[best].setdefault(slot, text)

    names = collections.defaultdict(lambda: collections.defaultdict(set))
    articles = collections.defaultdict(set)
    for key, rows in by_entry.items():
        i = owner.get(key)
        if i is None:
            continue
        if key in article:
            articles[i].add(article[key])
        for _, kind, lang, text in rows:
            names[i][text].add((key[0], kind, (lang or "").lower()))
    return names, intruders, articles


def is_english(lang):
    """⚠️ 'en' AND 'en-*', NEVER 'en*'. A prefix test on the bare string matches 'enm',
    which is MIDDLE ENGLISH, and this file used one at four places. Measured on the kept
    rows: 96 names over 35 rows were counted as English when a source had tagged them a
    dead language. 'honey' held mede and med, 'meat' held flesh and gos, 'half-and-half'
    held creme and 'verjuice' held verjus. Nobody writes a recipe in Middle English, and
    an English name is what the canonical picker picks and what english_names() returns
    to the sheet, so the wrong ones were reaching a reader.

    Every en* code in the store, measured: en 146,850 rows, en-gb 1,278, enm 912,
    en-ca 829, en-us 536. Four regional Englishes and one dead language, and the prefix
    could not tell them apart."""
    lang = (lang or "").lower()
    return lang == "en" or lang.startswith("en-")


def choose_canonical(entry, by_entry, stored_names):
    """⚠️ THE ANCHOR'S OWN ENGLISH LABEL WINS. No promotion off a binomial and no tiebreak
    on string length. An earlier build did both and picked 'ail' for garlic, because the
    French name is shorter, and 'wild garlic' for Allium sativum, which is a different
    plant. A binomial canonical is FLAGGED instead and the cook's name is picked by hand."""
    for key in sorted(entry["seed"]):
        for _, kind, lang, text in by_entry[key]:
            if kind in PRIMARY_KINDS and is_english(lang):
                return text, "the anchor's own English name"
    for key in sorted(entry["seed"]):
        name = stored_names.get((key[0], key[2]))
        if name:
            return name, "the anchor's own stored name, which is not in English"
    return None, "the anchor carries no name of its own"


def english_names(variations):
    """A name is English when a source states en or a regional en-*, or when it comes
    from wikipedia_redirect, which is enwiki and states no language at all.

    ⚠️ NOT A PREFIX TEST. See is_english."""
    return {text for text, tags in variations.items()
            if any(is_english(lang) or source == "wikipedia_redirect"
                   for source, _, lang in tags)}


def load_removals(path=HAND_REMOVALS):
    """Andy's hand removals, keyed on (anchor, id).

    ⚠️ (anchor, id) IS THE KEY BECAUSE THE CANONICAL NAME IS NOT UNIQUE. Measured over
    11,153 entries: zero (anchor, id) collisions, and 69 canonical names used by more
    than one entry. The id is the SOURCE'S OWN identifier, a Q-number or an OFF slug,
    not something this file invents.

    ⚠️ A ROW WITH NO REASON IS REJECTED, the same rule as ingredient_cuts.OVERRIDES."""
    if not os.path.exists(path):
        return {}, []
    with open(path, encoding="utf-8") as fh:
        rows = list(csv.DictReader(l for l in fh if not l.startswith("#")))
    removals, rejected = collections.defaultdict(list), []
    for row in rows:
        if not (row.get("reason") or "").strip():
            rejected.append(row)
            continue
        removals[(row["anchor"], str(row["id"]))].append(row)
    return dict(removals), rejected


def load_authored(path=AUTHORED_ROWS):
    """Rows Andy authored, the mirror of load_removals.

    ⚠️ AN AUTHORED ROW HAS NO ANCHOR AND THAT IS THE POINT. Every other row traces to a
    fetch. These exist because the thing exists and no source we hold has it: 'salt' is
    52 recipe lines and 32 dependents and had no row, so a line saying salt reached
    'table salt' or 'sea salt', which is the wrong specific rather than the general
    thing. anchor is left EMPTY rather than set to "authored", because a fourth source
    name would read as data and there is no fetch, no entry_id and nothing to
    re-derive behind it.

    ⚠️ A ROW WITH NO REASON IS REJECTED, the same rule as OVERRIDES and hand_removals.

    ⚠️ THE OPTIONAL seed COLUMN IS WHAT MAKES EXTRACTION POSSIBLE. Without it an authored
    row carries one name and no provenance, which is right for 'salt' and wrong for a
    member read out of a category: sherry taken off 'fortified wine' has four sources and
    fourteen languages behind its names, and dropping all of that to write the row would
    lose more than the row adds. See add_authored and seed_keys."""
    if not os.path.exists(path):
        return [], []
    with open(path, encoding="utf-8") as fh:
        raw = list(csv.DictReader(l for l in fh if not l.startswith("#")))
    good, rejected = [], []
    for row in raw:
        (good if (row.get("reason") or "").strip() else rejected).append(row)
    return good, rejected


def collect_variations(seed, by_entry, by_bucket):
    """Every name the seeded entries carry, each with the source, field and language that
    supplied it. Shared by add_overrides and add_authored.

    ⚠️ NOT AN ADMISSION ROUTE, AND THE DISTINCTION IS THE WHOLE POINT. Nothing here
    decides that a row should exist. A hand-written line already decided that, and this
    only answers "what is this thing called, and who says so". The seed is read as
    (source, dataset, entry_id) keys, so a name arrives with its provenance attached
    rather than as a bare string.

    Keys absent from by_entry are dropped rather than raising, because a seed can outlive
    an entry the vocabularies stopped shipping."""
    seed = {k for k in seed if k in by_entry}
    variations, buckets = collections.defaultdict(set), set()
    for key in seed:
        for norm, kind, lang, text in by_entry[key]:
            variations[text].add((key[0], kind, (lang or "").lower()))
            buckets.add(norm)
    for bucket in buckets:
        for s, d, e, kind, lang, text in by_bucket[bucket]:
            if (s, d, e) in seed:
                variations[text].add((s, kind, (lang or "").lower()))
    return variations


SEED_SOURCES = ("agrovoc", "off_taxonomy", "wikidata", "wikipedia_redirect", "wiktextract")


def seed_keys(spec, by_entry, by_bucket):
    """Read authored_rows.csv's seed column into (source, dataset, entry_id) keys.

    Three token forms, semicolon-separated:

      Q13228                       a Wikidata item
      off_taxonomy:en:heavy-cream  one named entry, in every dataset that carries it
      heavy cream                  a bucket, every entry any source files under the name

    ⚠️ THE BUCKET FORM IS THE BLUNT ONE AND IT IS NOT THE DEFAULT FOR A MEMBER. Measured
    while extracting 'cream': seeding 'heavy cream' as a bucket returns 168 names in 147
    languages, because en.wikipedia redirects 'Heavy cream' at the Cream article and the
    redirect sits in the same bucket, so the seed drags Q13228 and the whole parent
    concept onto the child row. 'single cream' does the same and adds 'light cream' on
    top. A member is seeded by ENTRY. The bucket form is kept because add_overrides needs
    it, and because it is right when the row IS the bucket.

    ⚠️ A SOURCE-AND-ID TOKEN EXPANDS ACROSS DATASETS ON PURPOSE. Open Food Facts is loaded
    twice under different field names, and 4,173 wiktextract entry ids appear in both the
    English and Chinese dumps. Those are copies of one entry, not two opinions, so the
    token takes all of them. See docs/measuring-the-premise.md, case 7.

    ⚠️ A SEED NAMES A CONCEPT AND USUALLY TAKES SEVERAL TOKENS TO DO IT. Measured on
    fortified wine, its 15 candidate names cover 5 concepts, so roughly three names per
    concept: Port, port, Porto, Port wine, Port Wine (DOC) and Vinho do Porto are one
    thing spelled six ways. Grouping the names is the reading job and it happens before
    this function, in a person's head. Semicolons are how the result is written down.

    ⚠️ THE PARENT SITS IN THE CHILD'S BUCKET, AND IT WILL FIRE ON EVERY EXTRACTION. This is
    not a run of unlucky cases, it is the shape of the data: a category and its member
    share names, which is WHY the member was on the category row in the first place, so
    the bucket that reaches the member also reaches the parent. Seeding by bucket pulls
    the parent's whole name set onto the child.

    Every case met so far, and the list is only going to grow:
      heavy cream    the bucket holds Q13228, cream, 168 names in 147 languages
      single cream   the same, plus light cream on top
      all NINE sausage members  every bucket holds Q131419, Sausage
      gnocchi, maultasche       both hold Q1854639, dumpling
      knödel                    holds Q5265534, the existing knedle row
      Ceylon cinnamon           holds Q28165, cinnamon, 253 names
      palatschinke              holds Q12200, crêpe
      tamari                    held Q3514675, which WAS already a row, and the row split

    THE GUARD IS TO SEED BY ENTRY, and build() reports any seed that names an entry
    another kept row is anchored on, so the trap is caught at build time rather than by
    reading the result. See the seed-collision check there."""
    keys = set()
    for token in (t.strip() for t in (spec or "").split(";")):
        if not token:
            continue
        if re.fullmatch(r"Q\d+", token):
            keys.add(("wikidata", "food_items_q2095", token))
        elif token.split(":")[0] in SEED_SOURCES and ":" in token:
            source, entry = token.split(":", 1)
            keys |= {(s, d, e) for (s, d, e) in by_entry
                     if s == source and e == entry}
        else:
            keys |= {(s, d, e) for s, d, e, *_ in by_bucket.get(norm_name(token), ())}
    return keys


def add_authored(rows, authored, subclass_count, by_entry, by_bucket):
    """Build the row dicts.

    ⚠️ AN UNSEEDED ROW KEEPS ITS EMPTY sources, so the 'only one source' flag cannot fire
    on a row that has none and the authored flag says the true thing instead.

    ⚠️ A SEEDED ROW IS A DIFFERENT CLAIM AND CARRIES A DIFFERENT RECORD. Extraction is
    where this matters. Reading 'fortified wine' and naming sherry as a member creates a
    row that no source made an entry, but the NAMES are not invented: Open Food Facts,
    Wikidata and AGROVOC all say sherry, and that provenance existed before the row did.
    So a seeded row's sources are computed from the tags rather than left blank, and the
    one-source flag keeps telling the truth about the names instead of lying by omission
    about all of them. What stays unsourced is the row's EXISTENCE, which is what the
    authored flag and the low-confidence floor already say.

    ⚠️ THE PLACEHOLDER 'authored' TAG IS EXCLUDED FROM sources DELIBERATELY. It marks the
    canonical as English without inventing a fetch, and the sheet renders source names
    through SOURCE_NAME, where the word would be a KeyError as well as a false claim."""
    for row in authored:
        name = row["name"].strip()
        sources = [s.strip() for s in (row.get("sources") or "").split(";") if s.strip()]
        seeded = collect_variations(seed_keys(row.get("seed"), by_entry, by_bucket),
                                    by_entry, by_bucket) if row.get("seed") else {}
        # The tag makes the name English to english_names() without inventing a source:
        # the language is a fact about the string, not a claim by anyone.
        variations = dict(seeded)
        variations.setdefault(name, set()).add(("authored", "label", "en"))
        if seeded:
            sources = sorted({s for tags in variations.values() for s, _, _ in tags}
                             - {"authored"})
        rows.append({
            "canonical": name, "how": "authored by hand, no source",
            "variations": variations, "n_variations": len(variations) - 1,
            "anchor": "", "id": row["id"].strip(), "kinds": [],
            "why": f"AUTHORED. {row['reason'].strip()}",
            "sources": sources, "seeded": bool(seeded),
            "languages": sorted({l for tags in variations.values()
                                 for _, _, l in tags if l}) or ["en"],
            "subclasses": subclass_count.get(row["id"].strip(), 0),
            "binomial": False, "rule2": False, "drink": False, "override": False,
            "dish": False,
            "authored": True, "added": (row.get("added") or "").strip(),
            "intruders": set(),      # nothing was absorbed, so nothing can intrude
            "articles": [],
            "dropped": [], "dead_dropped": [],
            "strength_a": {}, "strength_b": [],
        })
    return rows


def load_renames(path=HAND_RENAMES):
    """Andy's canonical renames, keyed on (anchor, id) like load_removals.

    ⚠️ A ROW WITH NO REASON IS REJECTED, the same rule as OVERRIDES, hand_removals and
    authored_rows."""
    if not os.path.exists(path):
        return {}, []
    with open(path, encoding="utf-8") as fh:
        raw = list(csv.DictReader(l for l in fh if not l.startswith("#")))
    good, rejected = {}, []
    for row in raw:
        if (row.get("reason") or "").strip() and (row.get("name") or "").strip():
            good[(row["anchor"].strip(), str(row["id"]).strip())] = row
        else:
            rejected.append(row)
    return good, rejected


def apply_renames(rows, renames):
    """Change a row's canonical to a name already sitting on it.

    ⚠️ 26 ROWS ANSWERED UNDER A NAME NOBODY TYPED, over 202 recipe lines. 'white sugar'
    carried 4 lines and held 'granulated sugar' at 35. 'corn starch' carried 0 and held
    'cornstarch' at 20. Most of the class is the American reading applied to rows written
    before that rule existed.

    ⚠️ NOTHING IS LOST AND THE CHANGE REVERSES BY NAME. The old canonical stays on the row
    as a variation marked 'derived', the same way strip_as_food records a stripped suffix,
    so a line using the old name still resolves to the same row.

    THREE REFUSALS, each reported rather than silently skipped:
      the new name is not already ON the row, which would put a string in the library
        that no source ever stated,
      the new name is already another KEPT row's canonical, which is a merge rather than
        a rename and needs a person to say which row keeps it,
      the key matches no row, which usually means a rule dropped the entry first."""
    owned = {norm_name(row["canonical"]) for row in rows if not row.get("cut_by")}
    done, refused = [], []
    for row in rows:
        rule = renames.get((row["anchor"], str(row["id"])))
        if not rule:
            continue
        new = rule["name"].strip()
        match = next((t for t in row["variations"] if norm_name(t) == norm_name(new)), None)
        if match is None:
            refused.append((rule, f"'{new}' is not a name on the row"))
            continue
        if norm_name(new) in owned - {norm_name(row["canonical"])}:
            refused.append((rule, f"'{new}' is already another row's canonical, so this "
                                  "is a merge rather than a rename"))
            continue
        row["variations"].setdefault(row["canonical"], set()).add(
            (row["anchor"] or "authored", DERIVED, "en"))
        row["renamed_from"] = row["canonical"]
        row["rename_reason"] = rule["reason"].strip()
        row["canonical"] = new
        row["n_variations"] = len(row["variations"]) - 1
        done.append(row)
    seen = {(r["anchor"], str(r["id"])) for r in done}
    for key, rule in renames.items():
        if key not in seen and not any(rule is x for x, _ in refused):
            refused.append((rule, "no row carries that (anchor, id)"))
    return done, refused


def apply_removals(rows, removals):
    """Mark entries 'hand' and trim variations. Nothing is deleted: a dropped entry moves
    to the cut sheet with the reason, and a trimmed name is recorded on the row it left."""
    seen = set()
    for row in rows:
        key = (row["anchor"], str(row["id"]))
        row["hand_reasons"], row["trimmed"] = [], []
        for rule in removals.get(key, ()):
            seen.add(key)
            action, reason = rule["action"], rule["reason"].strip()
            if action == "drop":
                row["hand_reasons"].append(reason)
            elif action == "drop_variation":
                name = (rule.get("variation") or "").strip()
                if name in row["variations"] and name != row["canonical"]:
                    del row["variations"][name]
                    row["trimmed"].append(f"{name} ({reason})")
            elif action == "trim_alias_only":
                gone = [t for t, tags in row["variations"].items()
                        if t != row["canonical"]
                        and all(k in ALIAS_KINDS for _, k, _ in tags)]
                for name in gone:
                    del row["variations"][name]
                if gone:
                    row["trimmed"].append(
                        f"{len(gone)} alias-only variation(s) ({reason}): "
                        + ", ".join(gone[:8]) + (" ..." if len(gone) > 8 else ""))
        row["n_variations"] = len(row["variations"]) - 1
    dangling = [r for key, rules in removals.items() if key not in seen for r in rules]
    return rows, dangling


# ─────────────────────────────────────────────────────────────────────────────────────
# Name resolution: which row answers to a name that two rows carry.
#
# Measured over the 11,153 rows before any of this ran: 1,301 (holder, name) pairs where
# a name on one row is ANOTHER row's canonical, and 490 of 2,997 recipe ingredient lines,
# 16.3%, hit a name that more than one row carries. That is the only group where the app
# gives a WRONG answer rather than no answer, so it is worth a rule where the rest is not.
#
# ⚠️ NOTHING IS DELETED HERE. A name that leaves a row is recorded ON the row it left with
#    the rule that moved it, the same way a cut row keeps its mark, so it reads back and
#    reverses by name.
#
# ⚠️ AND NO NAME STOPS RESOLVING, BY CONSTRUCTION. A pair only exists when another row
#    already claims that name as its CANONICAL, so every name taken off a holder still
#    answers, to the row that owns it. Measured across the whole pipeline: 1,262 of 2,997
#    recipe lines match a library name before and after, and 963 of them land on a
#    canonical before and after. Both figures are unchanged to the line.
# ─────────────────────────────────────────────────────────────────────────────────────
# ⚠️ AGROVOC WRITES A SYMBOL TWO WAYS AND THE FIRST DRAFT CAUGHT ONE. Some entries state
#    the bare form, AU for African Union and UN for United Nations. Others state it with
#    AGROVOC's own marker, 'Cu (symbol)' for copper and 'Al (symbol)' for aluminium. The
#    bare-only version missed Cu on honey, Al on garlic and Be on butter, which were three
#    of the five cases the rule was written for.
SYMBOL = re.compile(r"^[A-Z][a-z]?$|^[A-Z]{2,4}$|^[A-Z][a-z]? \(symbol\)$")


def drop_agrovoc_symbols(rows, by_entry):
    """A CHEMICAL SYMBOL IS NOT A NAME ANYONE WRITES, AND NEITHER IS AN ORGANIZATION'S
    INITIALS. 'honey' carried 'Cu', 'table salt' carried 'Na', 'grape' carried AU, EU,
    EEC and OAU, 'flour' carried UN, UNO, United Nations and United Nations Organization,
    'oyster' carried TMTD, thiram and thiuram.

    Same shape as drop_initialism_expansions and a different source. AGROVOC files a
    concept with its symbol as one English label and its name as another, so a bucket
    keyed on the symbol collides with any food whose name matches it in some language,
    and the whole entry arrives on the food row.

    THE ANCHOR CLAUSE, four conditions:
      every tag on the name is agrovoc, so it is sole evidence, and
      the name shares an AGROVOC entry with a symbol-shaped English label, and
      that label is one or two letters capitalized, or two to four capitals, and
      AGROVOC did not anchor the row, so the row is not the concept itself.

    MEASURED over the kept rows: 78 names leave 32 rows, and ZERO recipe lines reach any
    of them. They exist nowhere else afterwards, which is the same profile as the
    initialism cut, where 87 of 97 did and every one carried zero lines.

    ⚠️ WHAT IT REACHES AND WHAT IT DOES NOT, STATED EXACTLY, BECAUSE THE FIRST DRAFT WAS
    DESCRIBED WRONG. The bare-symbol-only version took 48 names off 14 rows and was
    reported as not reaching copper, nickel or Myanmar. Once the '(symbol)' form was
    added it DOES reach nickel on milk, aluminium on garlic and beryllium on butter,
    because those three sit on rows where AGROVOC is the sole evidence.

      REACHED   Cu (symbol), Al (symbol), Be (symbol), Ni (symbol), AU, EU, EEC, OAU, UN,
                UNO, TMTD, CAN, and the expansions on the same entry: African Union,
                European Union, United Nations, calcium ammonium nitrate, thiram, nickel,
                aluminium, beryllium.
      NOT       'copper' on honey, because Open Food Facts states it too, so it is not
                sole evidence and the guard keeps it.
      NOT       'Myanmar' on sesame oil, because AGROVOC's country entry states no symbol
                at all. Nothing on that entry is symbol-shaped, so there is no handle.

    ⚠️ NOBODY SHOULD WIDEN THIS TO REACH Myanmar. It arrives by a different route: the
    collision is on the food's own name in some language rather than on a symbol, so the
    entry that lands has nothing to catch it by. Widening to AGROVOC-only names in general
    would take real ingredient names with it, since AGROVOC alone carries some. Myanmar
    stays a reading job."""
    symbolic = set()
    for (source, _, entry_id), members in by_entry.items():
        if source != "agrovoc":
            continue
        english = [text for _, _, lang, text in members if is_english(lang)]
        if any(SYMBOL.match(text) for text in english):
            symbolic |= {norm_name(text) for text in english}

    gone = collections.Counter()
    for row in rows:
        if row["anchor"] == "agrovoc":
            continue                                  # the row IS the AGROVOC concept
        for text in list(row["variations"]):
            if text == row["canonical"] or norm_name(text) not in symbolic:
                continue
            if not all(source == "agrovoc" for source, _, _ in row["variations"][text]):
                continue                              # corroborated, so not sole evidence
            del row["variations"][text]
            row["dropped"].append(text)
            gone[row["canonical"]] += 1
        row["n_variations"] = len(row["variations"]) - 1
    return gone


# ⚠️ THE ONLY -ves PLURALS WHOSE SINGULAR IS NOT THE WORD MINUS ITS s. Every other word
#    ending in "ves" is an ordinary -s plural and must fall through: clove, glove, olive,
#    chive, dove, stove, sleeve. A blanket "ves" -> "f" rule shipped first and was wrong
#    about all of them, and about knife/life/wife too, because those take -fe not -f.
#    Membership, not shape, because no shape test separates cloves from wolves.
VES_SINGULAR = {
    "calves": "calf", "dwarves": "dwarf", "elves": "elf", "halves": "half",
    "hooves": "hoof", "knives": "knife", "leaves": "leaf", "lives": "life",
    "loaves": "loaf", "scarves": "scarf", "selves": "self", "sheaves": "sheaf",
    "shelves": "shelf", "thieves": "thief", "turves": "turf", "wharves": "wharf",
    "wives": "wife", "wolves": "wolf",
}


# ⚠️ THE PLURALS THE ss/us/is GUARD IN depluralize WRONGLY BLOCKS. A singular ending in -i
#    takes a plain -s, so its plural ends in "is" and is indistinguishable BY SHAPE from a
#    Greek or Latin singular. zucchinis and tennis both end "nis". chilis and iris both end
#    "lis". paninis and analysis both end in a consonant plus "is". Membership, not shape,
#    for the same reason VES_SINGULAR above is a list.
#
#    THE GUARD ITSELF STAYS, and the measurement is why. Over the whole 3,332-line recipe
#    corpus it blocks 26 distinct words and only chilis and zucchinis are real plurals.
#    Removing it would stem 'plus' to 'plu' on 97 lines, 'boneless' to 'boneles' on 18 and
#    'skinless' to 'skinles' on 14. Over the index the same holds: 1,418 keys end in "is"
#    and 74 of them look like plurals, so the blocked class is right about 95 percent of
#    the time and only the exceptions need naming.
#
#    Every entry below is either a word this bug was reported on or was generated from the
#    index (currently blocked, carries an English tag, and its stem is a real key) and then
#    read. The generated list also offered paris, propolis, sinapis and souris, which are
#    not plurals at all, so it is a candidate list and not the rule.
I_PLURAL = frozenset({
    "chilis", "chillis", "creminis", "kaseris", "kluskis", "litchis", "macaronis",
    "mueslis", "paninis", "pierogis", "pignolis", "pinolis", "raviolis", "rotis",
    "sushis", "tahinis", "ugalis", "uglis", "zucchinis",
})


def depluralize(key):
    """The English plural of a normalized name, or None. Deliberately small: three
    endings and a length floor, not a stemmer.

    ⚠️ IT IS WRONG ABOUT THREE NAMES IN 169 AND THEY ARE NAMED IN resolve_borrowed.

    ⚠️ -oes AND -ves WERE MISSING FROM THE FIRST DRAFT, WHICH MADE potatoes STEM TO
    'potatoe' AND leaves TO 'leave'. Worth almost nothing in moves, because a plural
    usually sits on the row its own singular names and rule 6 skips those by construction:
    41 pairs gain a correct stem and only TWO of them cross to a different row, both
    'Citrus hystrix' losing lime-leaf names, both at zero recipe lines. Fixed anyway,
    since a stemmer that is wrong about tomatoes is wrong wherever it is next used.

    ⚠️ AND THE -ves FIX THAT FOLLOWED IT, BECAUSE "ves" -> "f" WAS APPLIED TO EVERY WORD.
    cloves stemmed to 'clof', gloves to 'glof', olives to 'olif', chives to 'chif', and
    knives to 'knif' rather than 'knife'. Measured cost on the recipe corpus: 46 lines
    reading 'garlic cloves' reached no row, the single largest miss in the library, while
    'garlic clove' sits on Q28966859. A word not in VES_SINGULAR now falls through to the
    ordinary -s rule below, which is right for it. The suffix is matched on the LAST WORD
    so 'bay leaves' still reaches 'bay leaf'. Blast radius on a full rebuild: rows 11,217
    and kept 10,387 both unchanged, one move rule goes 255 to 260, and the five names that
    move are Conserves and Conſerves to conserve, Arbequina olives to arbequina olive, and
    Endives and endives to endive. All five land on a kept row.

    ⚠️ AND THE ss/us/is GUARD, WHICH BLOCKED EVERY PLURAL OF A SINGULAR ENDING IN -i.
    zucchinis, chilis, paninis, raviolis and macaronis all returned None, because the guard
    is a two-character suffix test and their plurals end in "is" exactly as analysis and
    basis do. The named exceptions in I_PLURAL run before the guard, on the last word, so
    'jalapeno chilis' reaches 'jalapeno chili'. Blast radius on a full rebuild: rows 11,217
    and kept 10,387 both unchanged, NO row's name set changes at all, and one move rule goes
    260 to 262, both of them Litchis swapping between the two duplicate litchi rows."""
    if key.endswith("ies") and len(key) > 4:
        return key[:-3] + "y"
    if key.endswith("ves") and len(key) > 4:
        head, _, last = key.rpartition(" ")
        single = VES_SINGULAR.get(last or key)
        if single:
            return (head + " " + single) if head else single
        # not a -f/-fe plural (cloves, olives): the ordinary -s rule below is correct
    if key.endswith("oes") and len(key) > 4:
        return key[:-2]
    if key.endswith(("ches", "shes", "sses", "xes", "zes")):
        return key[:-2]
    head, _, last = key.rpartition(" ")
    if (last or key) in I_PLURAL:
        return (head + " " + (last or key)[:-1]) if head else key[:-1]
    if key.endswith("s") and not key.endswith(("ss", "us", "is")) and len(key) > 3:
        return key[:-1]
    return None


AS_FOOD = " as food"
DERIVED = "derived"          # ⚠️ NOT A SOURCE FIELD. See strip_as_food.


def strip_as_food(rows):
    """Wikidata's "X as food" items name the food rather than the animal, and the suffix
    is Wikidata's disambiguator rather than a word a cook writes.

    ⚠️ THE SUFFIX COMES OFF ONLY WHERE THE STEM IS FREE. 106 rows carry it. On 19 the bare
    stem is ALREADY another row's canonical, and those 19 are exactly the rows where the
    distinction is doing work: 'lobster as food' against 'lobster', 'oyster as food'
    against 'oyster', 'goose as food' against 'goose', 'clam as food' against 'clam'.
    Wikidata holds the animal and the food as separate items and the library holds both,
    so a strip there would collide two real rows rather than rename one. Those 19 are a
    MERGE decision and they are left alone. 'egg as food' is one of them, and it carries
    46 of the 48 recipe lines the whole group reaches.

    The other 87 collide with nothing: blue marlin, crabgrass, geoduck, razor shell,
    blood vessel, lily bulb, shipworm, hagfish.

    ⚠️ THE STRIPPED NAME IS MARKED 'derived' RATHER THAN GIVEN A SOURCE FIELD, because no
    source wrote it. It is this file removing four characters, which is CURATED under
    docs/sourcing-tiers.md, and the variations cell has to say so or the row would read as
    though Wikidata had labelled it that."""
    owned = {norm_name(row["canonical"]) for row in rows}
    stems = collections.Counter(norm_name(row["canonical"][:-len(AS_FOOD)])
                                for row in rows
                                if row["canonical"].casefold().endswith(AS_FOOD))
    renamed = []
    for row in rows:
        if not row["canonical"].casefold().endswith(AS_FOOD):
            continue
        stem = row["canonical"][:-len(AS_FOOD)]
        if norm_name(stem) in owned or stems[norm_name(stem)] > 1:
            continue                                  # the 19. A merge, not a rename.
        row["variations"].setdefault(stem, set()).add((row["anchor"], DERIVED, "en"))
        row["renamed_from"] = row["canonical"]
        row["canonical"] = stem
        row["n_variations"] = len(row["variations"]) - 1
        renamed.append(row)
    return renamed


def drop_initialism_expansions(rows, by_entry):
    """A NAME WHOSE ONLY EVIDENCE IS THE EXPANSION OF AN INITIALISM IS NOT A NAME.

    English Wiktionary stores an initialism as an entry whose senses are its expansions,
    each linked back with alt_of. 'en:MPFE:noun#0' is meat, poultry, fish, eggs, so the
    join sees a bucket named 'eggs' pointing at an entry the meat anchor owns, and the
    meat row ends up carrying 'eggs'.

    ⚠️ THIS WAS READ AND RECORDED BEFORE IT WAS FIXED. reviewed.HEAD_TERMS marks five of
    them contaminated with the expansion spelled out: MPFE on eggs, SPG on salt, AP on
    all purpose flour, BEC on egg, VO on vegetable oil. Five of the fifty head terms in
    that sample, and the reading sat there unacted on because no rule reached the shape.

    ⚠️ WHY NAME RESOLUTION CANNOT REACH IT. resolve_borrowed moves a name only to a row
    that carries it as its CANONICAL. 'eggs' is plural and the authored row is 'egg', and
    norm_name does not stem, so the name meets no owner, wins against nothing and stays.
    The 20 recipe lines saying eggs have reached 'meat' through every pass so far.

    THE ANCHOR CLAUSE, and it is four conditions rather than one:
      the source is wiktextract, and
      the field is alt_of, NOT word, and
      the entry's word is an all-caps 2 to 6 letter string, and
      that tag is the ONLY evidence for the name on the row.

    ⚠️ alt_of AND NOT word, WHICH IS THE WHOLE PRECISION OF IT. The initialism itself is
    often a real abbreviation on the right row: AOVE is Spanish for extra virgin olive
    oil, RYR is red yeast rice, TSP is textured soy protein, WVO is waste vegetable oil.
    Dropping kind='word' as well would take 121 names instead of 97 and lose those four.
    The junk is entirely in the expansions.

    MEASURED over all 10,364 kept rows. 97 names leave 22 rows. 87 of the 97 exist
    nowhere else afterwards and every one of those 87 carries ZERO recipe lines. The
    three that carry lines all survive on another row, so no name stops resolving and no
    line loses its answer:

        'vegetable oil' leaves soursop      21 lines   expansion of VO
        'eggs'          leaves meat         20 lines   expansion of MPFE
        'egg'           leaves bacon        13 lines   expansion of BEC

    The rest is what an initialism list looks like: accounts payable and acute
    pancreatitis on all-purpose flour, sarong party girl and self-propelled gun on table
    salt, travelling salesman problem on textured soy protein, remote frame buffer on
    cream, and sixteen expansions of CS on Cabernet Sauvignon wine.
    """
    expansions = set()
    for (source, _, entry_id), members in by_entry.items():
        if source != "wiktextract":
            continue
        parts = entry_id.split(":")
        if len(parts) < 3 or not INITIALISM.match(parts[1]):
            continue
        for _, kind, lang, text in members:
            if kind == "alt_of":
                expansions.add((kind, (lang or "").lower(), text))

    gone = collections.Counter()
    for row in rows:
        row.setdefault("dropped", [])
        for text in list(row["variations"]):
            tags = row["variations"][text]
            if len(tags) != 1:
                continue                              # corroborated, so not sole evidence
            source, kind, lang = next(iter(tags))
            if source != "wiktextract" or (kind, lang, text) not in expansions:
                continue
            del row["variations"][text]
            row["dropped"].append(text)
            gone[row["canonical"]] += 1
    return gone


# ⚠️ A BINOMIAL IS NOT A NAME A COOK WRITES, SO LATIN IS NOT A MEMBER. Latin stays on
#    the row, because is_binomial rests entirely on the tag and all 56 flagged rows carry
#    it on the canonical. It is excluded HERE instead, where 778 Latin-only names over
#    555 rows would otherwise be counted as things to extract. The tag has two uses and
#    only one of them is a member.
MEMBER_EXCLUDED_LANGS = {"la"}


def member_names(row, owned):
    """The English names on a row that some source states as ITS OWN primary name and
    that no other row owns as a canonical. The extraction reading list.

    ⚠️ PRIMARY_CLAIM, NOT ANY FIELD THAT SOUNDS PRIMARY. An earlier sweep counted any
    name carried by a field called label, prefLabel, canonical_name, name, article_title
    or word, and reported 2,294 holders over 5,374 members. PRIMARY_CLAIM is
    source-qualified and excludes article_title and Wiktionary's word, which gives 1,698
    holders over 3,112 members. The difference is 42% and it was sizing the reading job.

    ⚠️ A NAME IS NOT A MEMBER, IT IS PART OF ONE. Measured twice: 'fortified wine' holds
    15 candidate names covering 5 concepts, and 'cream' holds 28 English names covering
    11. Port, port, Porto, Port wine, Port Wine (DOC) and Vinho do Porto are one thing
    spelled six ways. Grouping the names into concepts is the reading, and this function
    returns the names to be read rather than the rows to be written."""
    out = []
    for text, tags in row["variations"].items():
        if norm_name(text) == norm_name(row["canonical"]) or norm_name(text) in owned:
            continue
        if any(lang in MEMBER_EXCLUDED_LANGS for _, _, lang in tags):
            continue
        if any((source, kind) in PRIMARY_CLAIM and is_english(lang)
               for source, kind, lang in tags):
            out.append(text)
    return sorted(out)


def load_dead_languages(path=None):
    """vocab/dead-languages.tsv -> {code: (name, class)}. See the file's own header for
    how the list was built, and for why Latin is not in it."""
    path = path or os.path.join(VOCAB, "dead-languages.tsv")
    out = {}
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            if line.startswith("#") or not line.strip():
                continue
            code, name, kind, _ = line.rstrip("\n").split("\t")
            out[code] = (name, kind)
    return out


def drop_dead_language_names(rows, dead):
    """A NAME NOBODY ALIVE WRITES IS NOT A NAME. Middle English, Old Norse, Sanskrit,
    Classical Nahuatl and 102 reconstructed proto-languages.

    ⚠️ A FILTER, NOT A JUDGEMENT, AND IT FIXES A MEASUREMENT THAT WAS WRONG. 'flour' was
    measured at 16 members and ten of them were Middle English spellings of flour:
    fflour, fflowr, fleur, floure, flowr, flowre, flowyr, flor, flur, floor. Every holder
    in the library was showing more members than it has, so the reading job for
    extraction was sized against names no cook will ever type.

    THE ANCHOR CLAUSE, and it is one condition read strictly:
      EVERY tag on the name states a language, and every one of those languages is in
      vocab/dead-languages.tsv.

    ⚠️ 'EVERY' IS DOING THE WORK. A name any living source also states stays, whatever
    else is on it, so a Latin-and-English name or an Old-French-and-French name is
    untouched. The tag is removed from consideration, not the string, and the string
    only leaves when nothing living is left holding it.

    ⚠️ A NAME WITH NO LANGUAGE TAG AT ALL IS NOT TOUCHED. wikipedia_redirect states no
    language, and 42,762 name rows in the join carry a blank one. Reading blank as dead
    would empty the redirect source, which supplies 15,309 unique names.

    MEASURED over the kept rows. 1,166 names on 665 rows have only dead tags, of which
    the 150 non-Latin codes account for the cut. ⚠️ ZERO recipe lines reach ANY of them
    and ZERO rows lose every name, so nothing stops resolving and no row disappears.

    ⚠️ LATIN IS KEPT AND IT IS 3,420 OF THE 7,777 DEAD NAME ROWS. The binomial flag rests
    entirely on the Latin tag: 56 rows carry it and all 56 have a Latin tag on the
    canonical, so cutting 'la' would stop the flag rather than weaken it. What Latin
    should be excluded from is MEMBER COUNTING, where 778 Latin-only names over 555 rows
    inflate every holder, since a binomial is not a name a cook writes. Those two uses
    are separate and the flag keeps its evidence. See is_binomial."""
    gone = collections.Counter()
    for row in rows:
        row.setdefault("dead_dropped", [])
        for text in list(row["variations"]):
            if text == row["canonical"]:
                continue
            langs = {lang for _, _, lang in row["variations"][text] if lang}
            if not langs or not langs <= dead.keys():
                continue
            del row["variations"][text]
            row["dead_dropped"].append(text)
            gone[row["canonical"]] += 1
        row["n_variations"] = len(row["variations"]) - 1
    return gone


def strength_split(name):
    """(the strength words in a name, what is left of it). 'tomato paste' -> (['paste'], 'tomato')"""
    words = norm_name(name).split()
    return ([w for w in words if w in STRENGTH_WORDS],
            " ".join(w for w in words if w not in STRENGTH_WORDS))


def mark_strength(rows):
    """THE STRENGTH SHAPE. THE SAME INGREDIENT AT SEVERAL CONCENTRATIONS, IN TWO FORMS.

    A cook writing 'tamarind' and a cook writing 'tamarind concentrate' want the same
    fruit, and a spoonful of one is several spoonfuls of the other. That is not a naming
    collision and no existing mark sees it. The reader gets a plausible answer at the
    wrong ratio, which is worse to act on than an answer that is obviously wrong.

    FORM A, COLLAPSED. One row carries several strength words for one stem.
        tamarind paste holds paste, concentrate, extract, water, juice and liquid.
        tomato paste holds paste, concentrate and purée.
        broth holds bouillon, stock, bone stock and fish stock.
        cinnamon powder holds ground and powder, on 23 recipe lines.
      Measured: 148 rows, 98 recipe lines.

    FORM B, SCATTERED, AND IT IS THE LARGER ONE. Each strength has its OWN row and
    nothing links them. vanilla extract carries 37 recipe lines and sits beside vanilla
    powder, vanilla pod, vanilla sugar and pure vanilla extract as unrelated entries.
    lemon splits five ways, ginger six, lime five.
      Measured: 148 families over 402 rows, 128 recipe lines, 51 families with 3+ rows.
      ⚠️ THAT COUNT IS THE DETECTOR'S, NOT A READING. Hand-reading the top 18 by recipe
      line found 14 real families and 4 false ones ('food paste' beside 'raw food', 'ice
      cube' beside 'ice cream'), so the honest figure is on the order of 115.

    ⚠️ FORM B IS A RELATIONSHIP AND THE ROW MODEL CANNOT HOLD ONE. Every mark in this
    file describes a row on its own. Form B is a fact ABOUT TWO ROWS, and the only reason
    it can be written here at all is that the marker computes it across the whole set and
    copies the answer onto each member. That is the THIRD time the missing parent-child
    link has surfaced: the dish separation needed it, the category rows needed it, and now
    this. It is a schema gap rather than a data defect.

    ⚠️ IT HAS NOW SURFACED FIVE TIMES AND THE FIFTH IS THE CLEAREST. The tomato family
    varies on cultivar, cut and preservation at once, independently, and the sources have
    enumerated six of the roughly 800 combinations those three axes allow. 62 tomato rows
    exist and one carries a recipe line. 9 families vary on five axes and cheese is the
    largest at 163 rows against 33,600 combinations. All five surfacings, the arithmetic
    and what a model would have to hold are in docs/parent-child-gap.md.
    """
    families = collections.defaultdict(list)
    for i, row in enumerate(rows):
        if row.get("cut_by"):
            continue
        words, stem = strength_split(row["canonical"])
        if words and stem:
            families[stem].append((i, tuple(sorted(words))))

    for row in rows:
        row.setdefault("strength_a", {})
        row.setdefault("strength_b", [])

    a = b = 0
    for row in rows:
        seen = collections.defaultdict(set)
        for text in list(row["variations"]) + [row["canonical"]]:
            tags = row["variations"].get(text, {("", "", "en")})
            if not any(is_english(lang) for _, _, lang in tags):
                continue
            words, stem = strength_split(text)
            if words and stem:
                seen[stem] |= set(words)
        hit = {k: sorted(v) for k, v in seen.items() if len(v) > 1}
        if hit:
            row["strength_a"] = hit
            a += 1

    for stem, members in families.items():
        if len(members) < 2 or len({w for _, w in members}) < 2:
            continue                              # one row, or all the same strength
        names = [rows[i]["canonical"] for i, _ in members]
        for i, _ in members:
            rows[i]["strength_b"] = [c for c in names if c != rows[i]["canonical"]]
            b += 1
    return a, b


def resolve_borrowed(rows, superclasses, off_parents):
    """Take a name off a row when another row owns it as that row's canonical.

    RULE 1, A REDIRECT LOSES TO A CANONICAL AND WINS AGAINST NOTHING. A name supplied only
    by wikipedia_redirect leaves a row that another row claims as canonical. A redirect is
    a pointer to an article, not a claim that the target IS the thing, and en.wikipedia
    redirects 'Salmon' at 'Atlantic salmon' without saying they are the same fish.

    ⚠️ THE SECOND HALF OF THE RULE IS WHAT KEEPS THE SOURCE WORTH HAVING. Nothing is taken
    from a redirect that is the only source for a name. wikipedia_redirect supplies 15,309
    unique names over 947 recipe lines, and 14,773 of them over 608 lines are uncontested,
    so the rule keeps 96.5% of what the source uniquely gives and takes 3.5%. Every
    collapse term survives: Heeng, Windmill cookie, gochugaru, doubanjiang, guanciale,
    pekmez, za'atar, speculaas are all redirect-only and all uncontested.

    RULE 2, THE GENERAL TERM LEAVES THE SPECIFIC HOLDER. A name whose words are a strict
    subset of the holder's own canonical leaves it: 'salt' off 'sea salt', 'olive oil' off
    'extra virgin olive oil', 'milk' off "cow's milk". The specific row is not the general
    thing and it should not answer for it.

    Measured over the kept rows. Rule 1 alone takes 509 names, rule 2 alone takes 112,
    and 37 pairs are in BOTH, so the two together take 584 rather than 621. In the build,
    which resolves the cut rows as well and runs the rename first, rule 1 takes 500 and
    rule 2 takes the remaining 76.

    ⚠️ THE OVERLAP IS WHY RULE 2's HEADLINE NUMBER SHRINKS. Measured alone, rule 2 returns
    436 recipe lines and its largest single gain is 'olive oil' +117 off 'extra virgin
    olive oil'. Every one of those 117 is already returned by rule 1, which sees the same
    pair first because en.wikipedia is the only source that put 'olive oil' on that row.
    Behind rule 1, rule 2's gains are salt +156, sugar +44, egg +39, milk +22, oil +11,
    pepper +10, paprika +8, rice +4. ⚠️ SIX OF THE EIGHT ARE AUTHORED ROWS THAT DID NOT
    EXIST BEFORE STAGE 1. Rule 2 is worth almost nothing without authored_rows.csv,
    because there was no general row for the general term to go to.

    RULE 3, A SECOND PRIMARY NAME FROM ONE SOURCE IN ONE LANGUAGE LEAVES. A source gives a
    concept at most ONE primary name per language and all three keep that promise exactly,
    measured in sources.db rather than assumed: 0 of 250,765 AGROVOC (entry, language) pairs
    hold two prefLabels, 0 of 238,997 Wikidata pairs hold two labels, 0 of 61,565 OFF pairs
    hold two canonical_names. Zero exceptions in 551,327 pairs. So a row holding two is a
    row holding two concepts, and assign_ownership marks the later one.

    ⚠️ AND IT STILL ONLY LEAVES IF ANOTHER ROW CARRIES THE NAME, the same guard as rules 1
    and 2. Without that guard this rule could take a name out of the library, which the
    other two cannot do by construction.

    ⚠️ WHAT THIS DOES NOT REACH, AND IT IS 668 PAIRS. No rule touches a pair whose two
    names share no word, where no source is a redirect, and where both are the first primary
    name their source states in their language. 'peppercorn' holding 'pepper'
    has AGROVOC, Open Food Facts, Wikidata and Wiktionary behind it and they are not
    wrong. 'cabbage' holding 'water' has Open Food Facts behind it and is. Source count
    does not separate the two, so they are read by hand rather than ruled on.

    ⚠️ RULE 3 MARKS FAR MORE THAN IT MOVES, AND THE GUARD IS WHY. 14,936 second primary
    names are marked across 939 entries and 64 leave. 14,791 stay because nothing else in
    the library carries them, so removing one would take the name out altogether. They are
    named on the row in 'What I was unsure about' instead. The marks are a reading list, and
    deciding which of two entries keeps the row is a merge question, not a rule.

    RULE 4, AN AUTHORED ROW WINS THE NAME IT WAS AUTHORED FOR. Blast radius when it was
    added: one row. 'peppercorn' loses 'pepper' and 'Pepper' over 10 recipe lines, all ten
    the bare word. Rules 1 and 2 already clear salt, sugar, water, egg and oil.
    ⚠️ It overrides a reading reviewed.py records as the sources not being wrong, and that
    is deliberate. See the pepper entry there.

    RULE 5, A SEEDED AUTHORED ROW CLAIMS EVERY NAME IT WAS SEEDED WITH. ⚠️ WITHOUT IT
    EXTRACTION IS HALF DONE, and the case that showed it is 'wild garlic'. Extracting
    Allium ursinum off the garlic row moved the canonical and left ramsons, buckrams,
    bear leek, broad-leaved garlic and about 200 more names for the same plant sitting on
    Allium sativum, because no row owns those as a canonical and rules 1 to 4 only fire
    where one does.

    ⚠️ IT RESTS ON THE SEED BEING READ BY A PERSON, WHICH IS WHY IT IS SAFE. A seed names
    source entries, so the names it claims are the names those entries carry. Blast radius
    when it was added: 219 pairs, ZERO recipe lines, ZERO rows lost, and 213 of the 219
    are one plant leaving the wrong species.

    RULE 6, A PLURAL GOES TO THE ROW THAT OWNS ITS SINGULAR. 'eggs' carried 20 recipe
    lines on 'meat' and then on 'egg as food' and survived every pass, because norm_name
    does not stem and the authored row is 'egg', so the name met no owner and won against
    nothing. Second time the shape surfaced, after 'Egg yolks' on chicken egg yolk.

    ⚠️ THIS ONE MOVES THE NAME INSTEAD OF DELETING IT, AND IT IS THE ONLY RULE THAT DOES.
    The five above take a name the destination already carries AS ITS CANONICAL, so the
    name still answers by construction. A plural is not the destination's canonical.
    Deleting 'eggs' off 'egg as food' without putting it on 'egg' would take 20 recipe
    lines out of the library.

    ⚠️ MEASURED BEFORE APPLYING, AND THE GUARD CAME OUT OF THE MEASUREMENT. Unguarded, the
    stemmer fires on 191 pairs and is WRONG about nine, every one of them an English
    plural rule applied to a word that is not English: 'garos' to Garo, 'jambas' to Jamba,
    'maces' to mace, 'bigas' to biga, 'bolos' to bolo, 'Cremas' to Crema. Requiring the
    name to be English by english_names()'s own test, which counts wikipedia_redirect
    because enwiki states no language, gives 169 pairs and 31 recipe lines.

    ⚠️ THREE KNOWN FALSE POSITIVES, NAMED. 'Fideos' to Fideo, 'marrows' to marrow and
    'elks' to elk. Marrow is a vegetable and a bone, and an elk is a moose in Europe and
    a different animal in North America. All three carry ZERO recipe lines. They are
    recorded rather than special-cased, the same way is_binomial records American cheese.

    All nine pairs that carry recipe lines were read and all nine are right: eggs to egg
    at 20 lines, Egg yolks to egg yolk at 3, CARROTS off black carrot at 2, and Pork chops
    off pork ribs, bean sprouts, chicken wings, Button mushrooms and two spellings of
    chocolate chips at 1 each."""
    moved = collections.Counter()
    # ⚠️ A CUT ROW NEVER WINS A NAME, AND WITHOUT THIS GUARD TWO DID. 'Red Rome' moved off
    #    'Rome' and 'Bohnapfel' off 'Rheinischer Bohnapfel', both to rows the cultivar
    #    register then cut, which would have taken the two names out of the library
    #    altogether. Neither reaches a recipe line, so the cost was zero and the claim
    #    above was still false. The marks are computed here on the PRE-resolution row,
    #    which is safe in one direction: resolving only removes names, so it can push a
    #    row INTO a cut and never out of one.
    cut = [bool(apply_cuts(row, superclasses, off_parents)) for row in rows]
    canonical = collections.defaultdict(list)
    for i, row in enumerate(rows):
        if not cut[i]:
            canonical[norm_name(row["canonical"])].append(i)
    # ⚠️ RULE 5 READS A DIFFERENT INDEX, AND THAT IS THE WHOLE POINT. The four rules above
    #    only ever fire on a name ANOTHER ROW OWNS AS ITS CANONICAL, so a member's other
    #    names have no owner to go to and no rule reaches them. This index is every name a
    #    SEEDED authored row carries.
    seeded_claim = collections.defaultdict(list)
    for i, row in enumerate(rows):
        if cut[i] or not row["seeded"]:
            continue
        for text in row["variations"]:
            seeded_claim[norm_name(text)].append(i)

    for i, row in enumerate(rows):
        row.setdefault("resolved", [])
        head = set(norm_name(row["canonical"]).split())
        for text in list(row["variations"]):
            key = norm_name(text)
            if key == norm_name(row["canonical"]):
                continue
            owners = [j for j in canonical.get(key, ()) if j != i]
            if not owners:
                claim = [j for j in seeded_claim.get(key, ()) if j != i]
                if claim and not row["seeded"]:
                    del row["variations"][text]
                    row["resolved"].append((text, rows[claim[0]]["canonical"],
                                            "a seeded authored row claims the name"))
                    moved["a seeded authored row claims the name"] += 1
                    continue
                stem = depluralize(key)
                tags = row["variations"][text]
                if (stem and stem != key
                        and any(is_english(l) or s == "wikipedia_redirect"
                                for s, _, l in tags)):
                    holders = [j for j in canonical.get(stem, ()) if j != i]
                    if holders:
                        j = holders[0]
                        del row["variations"][text]
                        # ⚠️ MOVED, NOT DELETED, AND THIS IS THE DIFFERENCE FROM EVERY
                        #    RULE ABOVE. Rules 1 to 5 take a name the destination already
                        #    carries as its canonical, so the name still answers by
                        #    construction. A PLURAL IS NOT THE DESTINATION'S CANONICAL.
                        #    Deleting 'eggs' off 'egg as food' without putting it on 'egg'
                        #    would take 20 recipe lines out of the library altogether.
                        rows[j]["variations"].setdefault(text, set()).update(
                            (src, DERIVED, lang) for src, _, lang in tags)
                        row["resolved"].append((text, rows[j]["canonical"],
                                                "the plural of a name another row owns"))
                        moved["the plural of a name another row owns"] += 1
                continue                              # a redirect wins against nothing
            tags = row["variations"][text]
            if all(source == "wikipedia_redirect" for source, _, _ in tags):
                rule = "redirect loses to a canonical"
            elif set(key.split()) < head:
                rule = "the general term leaves the specific holder"
            elif text in row.get("intruders", ()):
                rule = "a second primary name from one source in one language"
            elif rows[owners[0]]["authored"]:
                rule = "an authored row wins the name it was authored for"
            else:
                continue
            del row["variations"][text]
            row["resolved"].append((text, rows[owners[0]]["canonical"], rule))
            moved[rule] += 1

    for row in rows:
        # ⚠️ RECOMPUTED, NOT LEFT STALE. A name that belongs to another row was never
        #    evidence for this one, so the source that supplied only that name stops
        #    counting here. Measured: 14 rows change source count and 4 fall to a single
        #    source (camembert de Normandie, wild carrot, pesto variants, Dutch Mimolette),
        #    which correctly moves all four into the one-source flag.
        row["n_variations"] = len(row["variations"]) - 1
        if row["authored"] and not row["seeded"]:
            # ⚠️ AN UNSEEDED AUTHORED ROW KEEPS ITS EMPTY sources, WHICH IS THE POINT OF
            #    add_authored. Recomputing from the tags would put the placeholder word
            #    'authored' into the source list, where it would read as a source name and
            #    would fire the "only one source says this exists" flag on a row that has
            #    none. A SEEDED row is recomputed like any other, minus that placeholder,
            #    because its names really do have sources and a name that left the row
            #    should stop counting as evidence for it.
            continue
        row["sources"] = sorted({s for tags in row["variations"].values()
                                 for s, _, _ in tags} - {"authored"})
        row["languages"] = sorted({l for tags in row["variations"].values()
                                   for _, _, l in tags if l})
    return moved


def apply_cuts(row, superclasses, off_parents):
    """⚠️ EVERY CUT NAMES AN ANCHOR. Read the anchor clause in ingredient_cuts.py before
    adding one: a cut phrased only as "single source and no variations" removes every
    override at once, because an override exists precisely because nothing corroborates
    the term."""
    marks = []
    if row.get("hand_reasons"):
        marks.append("hand")                          # ⚠️ Andy's call outranks every rule
    if (row["anchor"] == "wikidata"
            and set(superclasses.get(row["id"], [])) & set(CUTS.CULTIVAR_CLASSES)
            and len(row["sources"]) == 1 and row["n_variations"] == 0):
        marks.append("cultivar_register")
    # ⚠️ off_only USED TO FIRE HERE, on anchor == off_taxonomy and one source, and it
    #    took 3,549 rows. REVERSED, and the measurement that killed it is recorded in
    #    ingredient_cuts.DECLINED. It read "Open Food Facts is the only source" as
    #    "therefore this is label vocabulary", which is a claim about source coverage
    #    rather than about what a thing is. 56.7% of a read 60 were ordinary ingredients.
    #    ⚠️ NOTHING REPLACES IT YET, and that is deliberate. Read the label rows and put
    #    the verdicts in hand_removals.csv. A rule read back from evidence can come later.
    if (row["anchor"] == "off_taxonomy"
            and set(off_parents.get(row["id"], [])) & CUTS.OFF_FLAVOURING_PARENTS):
        marks.append("off_flavouring")
    return marks


def build_rows(join, src, kinds, superclasses, off_parents):
    by_entry, by_bucket = read_members(join)
    rule1, rule2, drinks, off_groups = pick_anchors(by_entry, by_bucket, kinds)

    stored_names = {}
    if src is not None:
        for table, source in (("wikidata_entry", "wikidata"),
                              ("off_taxonomy_entry", "off_taxonomy")):
            for entry_id, name in src.execute(f"SELECT entry_id, name FROM {table}"):
                stored_names[(source, entry_id)] = name

    entries = []
    for q in sorted(rule1 | rule2 | drinks):
        key = ("wikidata", "food_items_q2095", q)
        if q in rule1:
            why = "Wikidata kind is Ingredient or foodstuff"
        elif q in rule2:
            why = "Wikidata carries no kind, an OFF ingredient entry shares its name"
        else:
            why = ("Wikidata calls it a drink and an OFF ingredient entry carries the "
                   "same English name")
        entries.append({"anchor": "wikidata", "id": q, "seed": {key},
                        "buckets": {n for n, *_ in by_entry[key]}, "why": why})
    for group in off_groups:
        name = stored_names.get(("off_taxonomy", group[0][1]), "")
        if E_NUMBER.match(name.strip()) or name.strip().lower() in PROCESS_NAMES:
            continue                                  # additives and process words
        seed = {("off_taxonomy", d, e) for d, e in group}
        entries.append({"anchor": "off_taxonomy", "id": group[0][1], "seed": seed,
                        "buckets": {n for k in seed for n, *_ in by_entry[k]},
                        "why": "an Open Food Facts ingredients-taxonomy entry that "
                               "reaches no Wikidata item"})

    owned, intruders, articles = assign_ownership(entries, by_entry, by_bucket)
    subclass_count = collections.Counter()
    for parents in superclasses.values():
        for parent in parents:
            subclass_count[parent] += 1

    rows, dropped = [], collections.Counter()
    for i, entry in enumerate(entries):
        variations = owned.get(i)
        if not variations:
            continue
        item_kinds = sorted(kinds.get(entry["id"], {}).get("kinds", {}))
        fatal = [FATAL_KINDS[k] for k in item_kinds if k in FATAL_KINDS]
        if fatal:
            dropped[fatal[0] + ", even though Wikidata also calls it an ingredient"] += 1
            continue
        canonical, how = choose_canonical(entry, by_entry, stored_names)
        if canonical is None:
            best = [(len({s for s, _, _ in tags}), t) for t, tags in variations.items()
                    if any(k in PRIMARY_KINDS for _, k, _ in tags)]
            canonical = max(best)[1] if best else sorted(variations)[0]
        rows.append({
            "canonical": canonical, "how": how, "variations": variations,
            "n_variations": len(variations) - 1, "anchor": entry["anchor"],
            "id": entry["id"], "kinds": item_kinds, "why": entry["why"],
            "sources": sorted({s for tags in variations.values() for s, _, _ in tags}),
            "languages": sorted({l for tags in variations.values()
                                 for _, _, l in tags if l}),
            "subclasses": subclass_count.get(entry["id"], 0),
            "binomial": is_binomial(canonical, variations),
            "rule2": entry["why"].startswith("Wikidata carries no kind"),
            "drink": entry["why"].startswith("Wikidata calls it a drink"),
            "dish": bool(DISH_KINDS & set(item_kinds)) and INGREDIENT in item_kinds,
            "override": False, "authored": False, "seeded": False,
            # ⚠️ Names this row holds that are a SECOND primary name from one source in one
            #    language. See assign_ownership. resolve_borrowed decides what happens.
            "intruders": {t for t in intruders.get(i, ()) if t in variations},
            # ⚠️ THE en.wikipedia ARTICLES THIS ROW'S REDIRECT NAMES CAME FROM. Two or more
            #    means the row answers for two things en.wikipedia keeps apart. See
            #    the MEMBERS note in NOTES.
            "articles": sorted(articles.get(i, ())),
        })
    return rows, dropped, by_entry, by_bucket


def add_overrides(rows, by_entry, by_bucket, kinds, subclass_count):
    """The five hand-added terms. ⚠️ Report the list's size after every change: if it
    passes a few dozen the anchor rule is drawn in the wrong place."""
    for term, (failure, ident, reason) in CUTS.OVERRIDES.items():
        seed = {(s, d, e) for s, d, e, *_ in by_bucket.get(term, [])}
        if ident.startswith("Q"):
            seed.add(("wikidata", "food_items_q2095", ident))
        seed = {k for k in seed if k in by_entry}
        if not seed:
            continue
        variations = collect_variations(seed, by_entry, by_bucket)
        canonical = next((t for key in sorted(seed) for _, k, l, t in by_entry[key]
                          if k in PRIMARY_KINDS and is_english(l)),
                         term)
        anchor = "wikidata" if ident.startswith("Q") else "wiktextract"
        rows.append({
            "canonical": canonical, "how": "the anchor's own English name",
            "variations": dict(variations), "n_variations": len(variations) - 1,
            "anchor": anchor, "id": ident, "kinds": sorted(kinds.get(ident, {}).get("kinds", {})),
            "why": f"OVERRIDE ({failure}). {reason}",
            "sources": sorted({s for tags in variations.values() for s, _, _ in tags}),
            "languages": sorted({l for tags in variations.values() for _, _, l in tags if l}),
            "subclasses": subclass_count.get(ident, 0),
            "binomial": is_binomial(canonical, variations), "rule2": False,
            "drink": False, "override": True,
            "intruders": set(),      # hand-seeded from one bucket, so nothing can intrude
            "articles": [],
            "dropped": [], "dead_dropped": [],
            "strength_a": {}, "strength_b": [], "seeded": False,
            "dish": bool(DISH_KINDS & set(kinds.get(ident, {}).get("kinds", {})))
                    and INGREDIENT in kinds.get(ident, {}).get("kinds", {}),
            "authored": False,
        })
    return rows


def annotate(rows, superclasses, off_parents):
    """Borrowed names, the five checks, confidence, and the cut marks.

    ⚠️ CONFIDENCE IS NOT A QUALITY AXIS AND NO CUT USES IT. 614 low entries carry three or
    more sources and Allium sativum is one of them. It sorts reading order, nothing else."""
    fold = lambda s: s.casefold().strip()
    by_canonical = collections.defaultdict(list)
    for i, row in enumerate(rows):
        by_canonical[fold(row["canonical"])].append(i)

    for i, row in enumerate(rows):
        row["borrowed"] = [
            (text, [rows[j]["canonical"] for j in by_canonical[fold(text)] if j != i])
            for text in row["variations"]
            if fold(text) != fold(row["canonical"])
            and any(j != i for j in by_canonical.get(fold(text), []))]
        row["english"] = english_names(row["variations"])
        row["alias_only"] = [t for t, tags in row["variations"].items()
                             if t != row["canonical"]
                             and all(k in ALIAS_KINDS for _, k, _ in tags)]
        row["cut_by"] = apply_cuts(row, superclasses, off_parents)

        flags = []
        for reason in row.get("hand_reasons", ()):
            flags.append(f"REMOVED BY HAND: {reason}")
        for note in row.get("trimmed", ()):
            flags.append(f"VARIATIONS TRIMMED BY HAND: {note}")
        if row["override"]:
            flags.append("ADMITTED BY HAND. See 'Why it is in the list' for the reason.")
        if row["authored"] and not row["seeded"]:
            flags.append("AUTHORED BY HAND AND UNSOURCED. No source in the store has this "
                         "term. GENERATED under docs/sourcing-tiers.md until traced.")
        if row["seeded"]:
            flags.append("AUTHORED BY HAND, NAMES SEEDED. No source made this an entry, "
                         "so the ROW is GENERATED under docs/sourcing-tiers.md. Its names "
                         "and their provenance are read from the store and are not.")
        if row.get("rename_reason"):
            flags.append(f"RENAMED BY HAND from '{row['renamed_from']}'. "
                         + row["rename_reason"])
        elif row.get("renamed_from"):
            flags.append(f"RENAMED from '{row['renamed_from']}'. Wikidata's 'as food' "
                         "suffix is its own disambiguator, and no row claims the stem.")
        if len(row["articles"]) > 1:
            flags.append(
                f"answers for {len(row['articles'])} things en.wikipedia keeps apart, and "
                "each is a separate article whose redirects this row absorbed whole: "
                + ", ".join(row["articles"][:5])
                + (" ..." if len(row["articles"]) > 5 else ""))
        staying = sorted(row["intruders"] - {t for t, _, _ in row.get("resolved", ())}
                         - {row["canonical"]})
        if staying:
            flags.append(
                f"{len(staying)} name(s) here are a SECOND primary name from one source in "
                "one language, which means a second concept, and NOTHING ELSE IN THE LIBRARY "
                "CARRIES THEM so they stay rather than be lost: " + ", ".join(
                    repr(t) for t in staying[:4]))
        if row.get("strength_a"):
            flags.append(
                "STRENGTH SHAPE, FORM A. This row carries one thing at several "
                "concentrations, so a line matching it can be out by a multiple: "
                + "; ".join(f"{k} as {', '.join(v)}" for k, v in
                            list(row["strength_a"].items())[:2]))
        if row.get("strength_b"):
            flags.append(
                f"STRENGTH SHAPE, FORM B. {len(row['strength_b']) + 1} rows hold this "
                "ingredient at different strengths and NOTHING IN THE MODEL LINKS THEM: "
                + ", ".join(row["strength_b"][:5])
                + (" ..." if len(row["strength_b"]) > 5 else ""))
        if row.get("dropped"):
            flags.append(
                f"{len(row['dropped'])} name(s) DROPPED, each a Wiktionary initialism "
                "expansion or an AGROVOC symbol entry, carried by nothing else: "
                + ", ".join(repr(t) for t in sorted(row["dropped"])[:4])
                + (" ..." if len(row["dropped"]) > 4 else ""))
        if row.get("dead_dropped"):
            flags.append(
                f"{len(row['dead_dropped'])} name(s) DROPPED as dead-language only, "
                "stated by no living language: "
                + ", ".join(repr(t) for t in sorted(row["dead_dropped"])[:4])
                + (" ..." if len(row["dead_dropped"]) > 4 else ""))
        if row.get("resolved"):
            flags.append(f"{len(row['resolved'])} name(s) moved to the row that owns "
                         "them: " + "; ".join(f"'{t}' -> {o} ({why})"
                                              for t, o, why in row["resolved"][:3]))
        if row["borrowed"]:
            flags.append(f"holds {len(row['borrowed'])} name(s) another entry claims as "
                         "its own: " + "; ".join(
                             f"'{t}' belongs to {', '.join(o)}"
                             for t, o in row["borrowed"][:3]))
        if len(row["sources"]) == 1:
            # ⚠️ THIS IS A COVERAGE FACT, NOT A QUALITY ONE, and off_only was reversed for
            #    reading it as the latter. Open Food Facts alone carries cumin seeds.
            flags.append("only one source says this exists "
                         f"({SOURCE_NAME[row['sources'][0]]})")
        if row["binomial"]:
            flags.append("the canonical name is a scientific binomial, so a cook's name "
                         "has to be picked from the variations by hand")
        if row["rule2"]:
            flags.append("ADMITTED BY RULE 2, the weakest rule" + (
                f", and {row['subclasses']} items subclass it, which is what a CATEGORY "
                "looks like" if row["subclasses"] else ""))
        if row["n_variations"] >= 5:
            flags.append(f"{row['n_variations']} variations, and a third of buckets this "
                         "large measured wrong")
        if row["alias_only"]:
            flags.append(f"{len(row['alias_only'])} variation(s) rest only on a synonym, "
                         "alt_of or form field, which measured 20.8% wrong")
        if reviewed is not None:
            verdict = reviewed.lookup(row["canonical"])
            if verdict:
                flags.append(f"HAND-READ: {verdict}")
        row["flags"] = flags
        row["n_flags"] = len(flags)

        # ⚠️ An authored row is floored at low on purpose. Nothing corroborates it, and
        #    low sorts it to the top of the reading order where it stays visible.
        if (row["borrowed"] or len(row["sources"]) == 1 or row["authored"]
                or (row["rule2"] and row["subclasses"])):
            row["confidence"] = "low"
        elif len(flags) >= 2:
            row["confidence"] = "medium"
        elif len(flags) == 1:
            row["confidence"] = "medium" if row["n_variations"] >= 5 else "high"
        else:
            row["confidence"] = "high"

    # ⚠️ MEMBERS ARE COMPUTED LAST, AGAINST THE KEPT SET. A name that another row owns is
    #    not a member, and a CUT row does not own anything, so this has to run after the
    #    cut marks above rather than beside the variations. See member_names.
    owned = {norm_name(row["canonical"]) for row in rows if not row["cut_by"]}
    for row in rows:
        row["members"] = member_names(row, owned) if not row["cut_by"] else []
    return rows


# ─────────────────────────────────────────────────────────────────────────────────────
# The sheet. Facts left of judgements, a visible divider, two blank columns for the
# owner's call, and every header carrying its own caveat as a cell comment.
# ─────────────────────────────────────────────────────────────────────────────────────
CUT_RULE_TEXT = {
    "hand": "hand: removed by Andy, reason in 'What I was unsure about'",
    "cultivar_register": "cultivar register name: subclasses a Wikidata cultivar class, "
                         "one source, no variations",
    "off_flavouring": "OFF flavouring: parent is en:natural-flavouring or en:flavouring",
}
NOTES = {
 "members": "English names on this row that a source states as ITS OWN primary name and "
   "that no other row owns. The extraction reading list.\n\n"
   "⚠ A NAME IS NOT A MEMBER, IT IS PART OF ONE. 'fortified wine' holds 15 names covering 5 "
   "concepts, 'cream' holds 28 English names covering 11. Port, port, Porto, Port wine, Port "
   "Wine (DOC) and Vinho do Porto are one thing spelled six ways, so GROUP the names into "
   "concepts first and decide which concepts are ingredients second.\n\n"
   "⚠ LATIN IS EXCLUDED HERE AND KEPT ON THE ROW. A binomial is not a name a cook writes, and "
   "778 Latin-only names over 555 rows would otherwise read as things to extract. The binomial "
   "flag still needs the tag, so the tag stays.\n\n"
   "⚠ AND THE HEAD OF THIS LIST IS MOSTLY CONTAMINATION RATHER THAN MEMBERS. 'garlic' holds "
   "artificial intelligence and aluminium, 'butter' holds burdock and beryllium, 'sesame oil' "
   "holds Myanmar. Read them, record what they were, do not write rows for them.",
 "canonical": "The name a cook would say, VERBATIM from the anchor's own English label.\n\n"
   "⚠ NO PROMOTION AND NO TIEBREAK BY LENGTH. An earlier build picked 'ail' for garlic and "
   "'wild garlic' for Allium sativum. A binomial canonical is FLAGGED instead.",
 "status": "kept or CUT.\n\n⚠ NOTHING IS DELETED. A cut row is MARKED with the rule that "
   "removed it, the same way join_exclusion records an excluded label row, so a cut can be "
   "read back and reversed by name.",
 "cut_txt": "WHICH RULE removed it. Blank on a kept row.\n\nThe rules, their thresholds and "
   "their known false positives live in ingredient_cuts.py.",
 "n_variations": "How many other names resolve to this entry. Five or more is a flag, not a "
   "verdict: garlic genuinely has hundreds across the languages the store holds.",
 "vars": "Every variation as  name  [language | sources | field]. ENGLISH FIRST, then "
   "by how many sources carry the name.\n\n⚠ THE CELL IS CAPPED AT 60 AND 832 ROWS "
   "OVERFLOW IT. Alphabetical order hid 32.9% of every variation in the store and 35.9% "
   "of the English names on the rows that overflow. What is hidden now is the least "
   "corroborated tail, and the last line says how many and how many were English.\n\n"
   "⚠ LANGUAGE IS SHOWN, and it is what makes the hing case readable. 'hing' on ginger is Zhuang. 'hing' on asafoetida "
   "is English.\n\nTRANSLATIONS ARE NOT HERE. join.db excludes them by rule, because including "
   "them puts a pillow in the guanciale bucket.",
 "sources": "Which of the five sources contribute anything.",
 "languages": "Every language code appearing on a variation, ENGLISH FIRST.\n\n"
   "⚠ SORTED ALPHABETICALLY THIS COLUMN COULD NOT ANSWER THE ONE QUESTION IT GETS ASKED. "
   "milk holds 312 codes and its 'en' sat at position 67, so the first twenty read "
   "'aa, ab, aeb-arab, af, am' and stopped. 176 rows had an English name the column hid.",
 "kinds": "The Wikidata grouping. SEVERAL KINDS IS THE DUAL NATURE, NOT AN ERROR.\n\n"
   "⚠ 'Cultivar or plant variety' matched zero of the 28,630 items, so the cultivar "
   "exclusion never fired. See vocab/README.md.",
 "anchor": "Which source made this an entry. Only Wikidata and OFF can anchor, plus the "
   "hand-added overrides.",
 "rule2": "TRUE for the 487 entries admitted by rule 2, the weakest rule.",
 "drink": "TRUE for the 22 entries admitted by rule 4, the drinks rule.\n\n"
   "⚠ A NARROWING, NOT A RELAXATION. Wikidata calls these drinks and not ingredients, so "
   "the kind filter excluded them and wine, coffee and tea had no row at all. The loose "
   "version, any OFF ingredient entry sharing a name bucket, admits 65 at 60%. Requiring "
   "the two ENGLISH names to be EQUAL admits 22 and is wrong on none of them.\n\n"
   "⚠ WHAT THE LOOSE VERSION LET IN WAS A CROSS-LANGUAGE HOMOGRAPH EVERY TIME, the third "
   "time that has been the failure here after 'ni' for nickel and 'gula' for sugar. "
   "'latte' matched OFF milk, 'Uva' matched grape, 'weak coffee' matched WATER on a "
   "bucket carrying 33 recipe lines.\n\nAll 65 are read in reviewed.DRINKS_SAMPLE, "
   "including the 19 this rule drops because their OFF match is a parent rather than the "
   "same thing. Appellations stay excluded. See build_library.drinks_rule.",
 "authored": "TRUE for rows Andy wrote, which no source in the store has.\n\n"
   "⚠ THESE ARE THE ONLY ROWS WITH NO ANCHOR. Every other row traces to a fetch. These "
   "exist because the thing exists and nothing we hold names it: 'salt' is 52 recipe "
   "lines and 32 dependents and had no row, so a line saying salt reached 'table salt' "
   "or 'sea salt', the wrong specific rather than the general thing.\n\n"
   "⚠ GENERATED under docs/sourcing-tiers.md until traced. Floored at low confidence "
   "because nothing corroborates them. The reason and the measurement behind each one "
   "are in authored_rows.csv.",
 "dish": "TRUE for the 653 rows Wikidata calls BOTH a dish or a cuisine AND an "
   "ingredient. Sort on it to read them.\n\n"
   "⚠ A MARK, NOT A VERDICT, AND NOT A REMOVAL QUEUE. Reading 40 at random found 16 "
   "dishes and 24 ingredients, 40% dishes, 95% CI [26%, 55%]. Moving all 653 would take "
   "chocolate, shrimp, honey, ham, soy sauce, oyster, carrot and rolled oats out of the "
   "library, and would cost 46 recipe lines.\n\n"
   "⚠ THE 'X as food' ROWS ARE NOT DISHES. pike as food, squid as food, razor shell as "
   "food, gar as food and dog cockle as food are Wikidata separating the ANIMAL from the "
   "FOOD, which is the distinction an ingredient library wants. 52 rows. Do not read them "
   "as removal candidates.\n\n"
   "⚠ Six discriminators were measured and none is usable, so nobody should propose one "
   "from memory. The best is a coin flip. The numbers are in ingredient_cuts.DECLINED "
   "under 'dish_separation'.",
 "confidence": "low / medium / high.\n\n⚠ CONFIDENCE IS NOT A QUALITY AXIS AND NO CUT USES IT. "
   "614 low entries carry three or more sources and Allium sativum is one of them.",
 "flags": "What I was unsure about, one per line. A line beginning HAND-READ is a verdict "
   "from reviewed.py, which is evidence rather than inference.",
 "n_flags": "How many checks fired. Sortable.",
 "subclasses": "How many Wikidata items name this one as a superclass.\n\n⚠ RANKS CATEGORIES, "
   "CANNOT DECIDE THEM. cheese has 86 children in OFF's tree, fruit 78, honey 44, rice 34.",
 "why": "Which admission rule let it in, or the recorded reason for a hand-added override.",
 "how": "How the canonical name was chosen.",
 "resolved": "Names this row USED TO carry that another row owns as its canonical, with "
   "the rule that moved each one.\n\n⚠ NOTHING IS LOST. A name only appears here when "
   "another row already claims it as its OWN canonical, so it still resolves, to that row "
   "instead of this one. Measured: 1,262 of 2,997 recipe lines match a library name before "
   "and after, unchanged to the line, while lines hitting a name TWO rows carry fall from "
   "490 to 266.\n\nThe two rules are in build_library.resolve_borrowed.",
 "articles": "The en.wikipedia ARTICLES whose redirects this row absorbed, shown only where "
   "there is more than one.\n\n⚠ A wikipedia_redirect entry is one article plus every name "
   "redirecting to it, so two articles on one row means the row answers for two things "
   "en.wikipedia keeps apart. fortified wine holds Fortified wine, Port wine and Sherry. "
   "dumpling holds ten, Gnocchi and Knedle and Knodel among them.\n\n⚠ THIS COSTS NOTHING "
   "AND WAS ALREADY IN join.db. 12,055 redirect entries, every one carrying its article "
   "title, and build_library read the field and ignored what it meant.\n\n218 rows are "
   "flagged over 281 recipe lines, 57.3% of them subclassed by something against a sheet "
   "baseline of 11.4%. 60 of the 218 carry NO second primary name, so the two signals are "
   "not the same check: peppercorn merging Pebre at 32 recipe lines is one of them.\n\n"
   "⚠ THE ROW COUNT AND THE LINE COUNT ARE BOTH COMPUTED FROM OWNERSHIP. A first pass "
   "mapped variation TEXT back to redirect entries instead and reported 340 rows over 168 "
   "lines. It double-counted a name several articles redirect, and missed rows whose "
   "article title is not itself a variation. Ownership is the truth: 218 and 281.",
 "strength_b": "OTHER ROWS holding this same ingredient at a different concentration.\n\n"
   "⚠ THE STRENGTH SHAPE, FORM B, AND IT IS A RELATIONSHIP RATHER THAN A PROPERTY. "
   "'vanilla extract' carries 37 recipe lines and sits beside 'vanilla powder' as an "
   "unrelated entry. lemon splits five ways, ginger six, lime five. A cook writing the "
   "bare word wants one of them and the model cannot say which, because nothing here "
   "records that they are one ingredient.\n\n148 families over 402 rows, 128 recipe "
   "lines, 51 families with three or more rows. ⚠ THAT IS THE DETECTOR'S COUNT. Reading "
   "the top 18 by line found 14 real and 4 false, so the honest figure is nearer 115.\n\n"
   "Form A, the collapsed version, is in 'What I was unsure about' instead, because it "
   "is a property of one row. See build_library.mark_strength.",
 "__blank": "Yours. Nothing is written here.", "__blank2": "Yours. Nothing is written here.",
 "__div1": "LEFT is copied verbatim from a source. RIGHT is my judgement.",
 "__div2": "RIGHT is yours.",
}
COLUMNS = [
 ("Ingredient", "canonical", 26), ("Kept or cut", "status", 10),
 ("Removed by which rule", "cut_txt", 44), ("Variations", "n_variations", 9),
 ("Every variation, with language, source and field", "vars", 70),
 ("Sources", "sources", 22), ("Languages", "languages", 20),
 ("What kind of thing it is", "kinds", 26), ("Anchored on", "anchor", 22),
 ("Admitted by rule 2", "rule2", 11), ("Admitted by the drinks rule", "drink", 12),
 ("Dish as well as ingredient", "dish", 12),
 ("Authored, not sourced", "authored", 12),
 ("Names moved to their owner", "resolved", 46),
 ("Wikipedia articles it answers for", "articles", 34),
 ("Same thing at another strength", "strength_b", 40),
 ("Member names to read for extraction", "members", 46),
 (">>> JUDGEMENTS BELOW >>>", "__div1", 4),
 ("Confidence", "confidence", 11), ("What I was unsure about", "flags", 60),
 ("Checks that fired", "n_flags", 9), ("Things that subclass it", "subclasses", 11),
 ("Why it is in the list", "why", 42), ("How the name was chosen", "how", 30),
 (">>> YOUR CALL BELOW >>>", "__div2", 4), ("My call", "__blank", 16),
 ("My note", "__blank2", 36),
]


def write_sheet(rows, path):
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    fills = {k: PatternFill("solid", fgColor=v) for k, v in {
        "header": "1F3864", "divider": "595959", "fact": "DDEBF7", "judge": "FCE4D6",
        "yours": "E2EFDA", "low": "FFC7CE", "medium": "FFF2CC", "high": "E2EFDA",
        "cut": "D9D9D9", "override": "D9D2E9"}.items()}
    edge = Side(style="thin", color="BFBFBF")
    border = Border(left=edge, right=edge, top=edge, bottom=edge)
    order = {"low": 0, "medium": 1, "high": 2}

    def variation_text(row, cap=60):
        """⚠️ ORDER BEFORE TRUNCATING. THE OLD CAP DROPPED THE END OF THE ALPHABET.

        Measured over 11,153 rows: 832 hold more than 60 variations and the alphabetical
        cap hid 65,766 of 199,728 variation rows, 32.9% of everything the store knows.
        Worse than the volume is which rows it fell on. All 832 are flagged "5 or more
        variations, and a third of buckets this large measured wrong", so the truncation
        landed exactly on the rows most worth reading, and it hid 5,816 of their 16,197
        ENGLISH names, 35.9%, on 454 of the 832.

        milk is the worked case. 523 variations, 40 of them English, and the visible cell
        opened with Abe', abe', akeffay, akʷfay, amata.

        English first, then by how many sources carry the name, then alphabetical. What
        falls off the end is now the least corroborated tail rather than the letters after
        the sixtieth."""
        items = []
        for text, tags in row["variations"].items():
            if text == row["canonical"]:
                continue
            langs = sorted({l for _, _, l in tags if l})
            sources = sorted({SOURCE_NAME[s] for s, _, _ in tags})
            items.append(((text not in row["english"], -len(sources), text.casefold()),
                          "{}   [{} | {} | {}]".format(
                              text, ", ".join(langs) if langs else "no language stated",
                              ", ".join(sources),
                              ", ".join(sorted({FIELD_NAME.get(k, k) for _, k, _ in tags})))))
        items.sort(key=lambda kv: kv[0])
        if len(items) <= cap:
            return "\n".join(text for _, text in items)
        hidden = items[cap:]
        n_en = sum(1 for key, _ in hidden if not key[0])
        return "\n".join([text for _, text in items[:cap]] + [
            f"... and {len(hidden)} more, {n_en} of them English. English names and the "
            "best-corroborated come first, so what is hidden is the least corroborated "
            "tail."])

    def language_text(row, cap=20):
        """⚠️ ENGLISH FIRST, BECAUSE THE QUESTION THIS COLUMN GETS ASKED IS WHETHER THE ROW
        HAS AN ENGLISH NAME AT ALL, and sorted alphabetically it could not answer.

        1,783 rows carry more than 20 language codes and the first-20 cut hid 57,589 of
        them. milk holds 312 and its 'en' sat at alphabetical position 67, so the column
        read aa, ab, aeb-arab, af, am and stopped. On 176 rows the code was present and
        invisible.

        ⚠️ Every row with an English name carries an 'en*' code, all 9,668 of them, so the
        column can answer the question once 'en' is put where it can be seen. The
        wikipedia_redirect rows that state no language of their own are all covered by an
        en-tagged name from another source."""
        first = [l for l in row["languages"] if is_english(l)]
        rest = [l for l in row["languages"] if not is_english(l)]
        shown = (first + rest)[:cap]
        extra = len(row["languages"]) - len(shown)
        return ", ".join(shown) + (f", ... and {extra} more" if extra else "")

    def sheet(ws, data, banner):
        divider = next(i for i, (_, k, _) in enumerate(COLUMNS, 1) if k == "__div1")
        for i, (head, key, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(1, i)
            cell.value = (banner if i == 1 else "DERIVED, my judgement"
                          if key == "confidence" else "YOURS" if head == "My call" else "")
            cell.fill = fills["fact"] if i < divider else (
                fills["yours"] if key in ("__blank", "__blank2", "__div2") else fills["judge"])
            cell.font = Font(bold=True, size=9, color="333333")
            cell.border = border
        for i, (head, key, width) in enumerate(COLUMNS, 1):
            cell = ws.cell(2, i)
            cell.value = head
            cell.fill = fills["divider"] if key.startswith("__div") else fills["header"]
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if key in NOTES:
                note = NOTES[key]
                comment = Comment(head.upper() + "\n\n" + note, "build_library.py")
                comment.width = 350
                comment.height = max(140, 32 + 13 * (len(note) // 50 + note.count("\n") * 2))
                cell.comment = comment
            ws.column_dimensions[get_column_letter(i)].width = width
        for r, row in enumerate(data, 3):
            values = {
                "canonical": row["canonical"], "status": "CUT" if row["cut_by"] else "kept",
                "cut_txt": "\n".join(CUT_RULE_TEXT[c] for c in row["cut_by"]),
                "n_variations": row["n_variations"], "vars": variation_text(row),
                "sources": ", ".join(SOURCE_NAME[s] for s in row["sources"]),
                "languages": language_text(row),
                "kinds": ", ".join(row["kinds"]) or (
                    "(authored, so no source classified it)" if row["authored"] else
                    "(OFF entry, Wikidata never classified it)"
                    if row["anchor"] == "off_taxonomy" else "(Wikidata carries no kind)"),
                "anchor": (f"authored by hand, no source  {row['id']}" if row["authored"]
                           else f"{SOURCE_NAME[row['anchor']]}  {row['id']}"),
                "rule2": "yes" if row["rule2"] else "",
                "drink": "yes" if row["drink"] else "",
                "dish": "yes" if row["dish"] else "",
                "authored": "yes" if row["authored"] else "",
                "resolved": "\n".join(f"{t}  ->  {o}   [{why}]"
                                      for t, o, why in row.get("resolved", ())),
                "articles": "\n".join(row["articles"]) if len(row["articles"]) > 1 else "",
                "strength_b": "\n".join(row.get("strength_b", ())),
                "members": "\n".join(row.get("members", ())),
                "confidence": row["confidence"],
                "flags": "\n".join(row["flags"]), "n_flags": row["n_flags"],
                "subclasses": row["subclasses"] or "", "why": row["why"], "how": row["how"]}
            for i, (head, key, width) in enumerate(COLUMNS, 1):
                cell = ws.cell(r, i)
                if key.startswith("__div"):
                    cell.fill, cell.border = fills["divider"], border
                    continue
                if key.startswith("__blank"):
                    cell.fill, cell.border = fills["yours"], border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    continue
                cell.value = values.get(key, "")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border, cell.font = border, Font(size=10)
                if key == "status":
                    cell.fill = fills["cut"] if row["cut_by"] else fills["high"]
                    cell.font = Font(size=10, bold=True)
                elif key == "confidence":
                    cell.fill = fills[row["confidence"]]
                    cell.font = Font(size=10, bold=True)
                elif row["override"]:
                    cell.fill = fills["override"]
        ws.freeze_panes = "B3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(COLUMNS))}{len(data) + 2}"
        ws.row_dimensions[1].height = 13
        ws.row_dimensions[2].height = 56

    kept = sorted((r for r in rows if not r["cut_by"]),
                  key=lambda r: (order[r["confidence"]], -r["n_flags"], -r["n_variations"],
                                 r["canonical"].casefold()))
    cut = sorted((r for r in rows if r["cut_by"]),
                 key=lambda r: (order[r["confidence"]], -r["n_flags"], -r["n_variations"],
                                r["canonical"].casefold()))
    rule2 = sorted((r for r in rows if r["rule2"]),
                   key=lambda r: (-r["subclasses"], -r["n_variations"],
                                  r["canonical"].casefold()))
    book = openpyxl.Workbook()
    sheet(book.active, kept, f"THE LIST, {len(kept):,} kept rows")
    book.active.title = "ingredients (kept)"
    sheet(book.create_sheet("cut, with the rule"), cut,
          f"CUT, {len(cut):,} rows. NOTHING DELETED. Filter 'Removed by which rule'.")
    sheet(book.create_sheet("rule 2, decide by hand"), rule2,
          f"THE {len(rule2)} RULE-2 ENTRIES, sorted by subclass count. Nothing removed.")
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    book.save(path)
    return len(kept), len(cut), len(rule2)


def guard_unharvested(path, removals):
    """⚠️ REFUSE TO OVERWRITE A SHEET CARRYING MARKS THE CSV HAS NOT SEEN.

    This is the whole reason marking in the spreadsheet is safe to recommend. Mark forty
    rows, forget to harvest, rebuild, and the reading would be gone. That is the exact
    failure that moved the overrides out of a scratchpad, so the build stops instead."""
    if not os.path.exists(path):
        return
    try:
        import openpyxl
        import harvest_marks
    except ImportError:
        return
    book = openpyxl.load_workbook(path, read_only=True)
    unseen, heads = [], None
    for ws in book:
        for i, values in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 2:
                heads = {str(v): j for j, v in enumerate(values) if v}
                continue
            if i < 3 or heads is None:
                continue
            def cell(name):
                j = heads.get(name)
                return values[j] if j is not None and j < len(values) else None
            call = cell("My call")
            if not (call and str(call).strip()):
                continue
            action, extra = harvest_marks.parse_call(str(call))
            if action in (None, "keep"):
                continue
            parts = str(cell("Anchored on") or "").split()
            anchor = harvest_marks.SOURCE_KEY.get(parts[0]) if parts else None
            if anchor is None:
                continue
            key = (anchor, " ".join(parts[1:]))
            if not any(r["action"] == action and (r.get("variation") or "") == extra
                       for r in removals.get(key, ())):
                unseen.append((cell("Ingredient"), call))
    if unseen:
        listed = "\n".join(f"      {n!r} marked {c!r}" for n, c in unseen[:12])
        more = f"\n      ... and {len(unseen) - 12} more" if len(unseen) > 12 else ""
        raise SystemExit(
            f"⚠️  REFUSING TO OVERWRITE {path}.\n"
            f"  It carries {len(unseen)} mark(s) that hand_removals.csv has not recorded, "
            "and rebuilding would discard them:\n"
            f"{listed}{more}\n\n"
            "  Run this first, then rebuild:\n"
            "      python3.13 harvest_marks.py")


def build(join_db=JOIN_DB, sources_db=SOURCES_DB, out=OUT, verbose=True):
    def say(*a):
        if verbose:
            print(*a, flush=True)

    if not os.path.exists(join_db):
        raise SystemExit(f"{join_db} not found. Run build_join.py first.")
    join = sqlite3.connect(f"file:{join_db}?mode=ro", uri=True)
    # ⚠️ sources.db IS REQUIRED, AND AN EARLIER DRAFT MADE IT OPTIONAL. That draft ran
    #    happily without it and produced 11,769 entries instead of 11,153, because the
    #    E-number and process-word filters read OFF's stored entry names and silently
    #    matched nothing, letting 614 additives and 2 process words through. Canonical
    #    names diverged too. A build that runs and quietly returns a different list is
    #    worse than one that stops, so it stops.
    if not os.path.exists(sources_db):
        raise SystemExit(
            f"{sources_db} not found, and it is REQUIRED.\n"
            "  It supplies OFF's stored entry names, which the E-number and process-word\n"
            "  filters need, and the English glosses. Without it this script would return\n"
            "  11,769 entries instead of 11,153 and would not say so.\n"
            "  Rebuild it with fetch_sources.py then build_sources_db.py, or restore it\n"
            "  from a backup. vocab/off-taxonomy-tree.json is a committed copy of the OFF\n"
            "  tree for reading, NOT a substitute for the corpus.")
    src = sqlite3.connect(f"file:{sources_db}?mode=ro", uri=True)
    src.execute("PRAGMA query_only = ON")

    kinds, superclasses, committed_tree = load_vocab()
    say(f"vocab loaded: {len(kinds):,} kinds, {len(superclasses):,} superclass rows")
    off_parents = off_tree(src) or committed_tree
    say(f"OFF tree: {len(off_parents):,} entries with a parent, rebuilt from sources.db")

    rows, dropped, by_entry, by_bucket = build_rows(join, src, kinds, superclasses,
                                                    off_parents)
    subclass_count = collections.Counter()
    for parents in superclasses.values():
        for parent in parents:
            subclass_count[parent] += 1
    rows = add_overrides(rows, by_entry, by_bucket, kinds, subclass_count)
    # ⚠️ AFTER the overrides and BEFORE the removals, so a hand removal can target an
    #    authored row the same way it targets any other. Its key is ("", <id>).
    authored, authored_rejected = load_authored()
    rows = add_authored(rows, authored, subclass_count, by_entry, by_bucket)

    removals, rejected = load_removals()
    rows, dangling = apply_removals(rows, removals)
    # ⚠️ AFTER the removals and BEFORE strip_as_food, for the reason strip_as_food gives:
    #    a rename changes the canonical set the precedence rules read, so every rename has
    #    to land before resolution runs.
    renames, rename_rejected = load_renames()
    renamed_by_hand, rename_refused = apply_renames(rows, renames)

    # ⚠️ THE RENAME RUNS FIRST AND THE PRECEDENCE RULES SECOND, because a rename changes
    #    the canonical set the precedence rules read. Measured both orders: renaming first
    #    costs 5 of rule 1's 509 names and 0 of rule 2's, and stripping 'as food' ADDS 7
    #    borrowed pairs, because a stem that was hidden behind a suffix now collides with
    #    names other rows hold. Precedence has to run after that or it misses those 7.
    for row in rename_rejected:
        say(f"  ⚠️  REJECTED, no reason or no name: rename {row.get('anchor')} "
            f"{row.get('id')}. A rename without a reason is not applied.")
    say(f"  renamed by hand: {len(renamed_by_hand):5,d}   "
        + ", ".join(f"{r['renamed_from']} -> {r['canonical']}"
                    for r in renamed_by_hand[:4])
        + (" ..." if len(renamed_by_hand) > 4 else ""))
    for rule, why in rename_refused:
        say(f"  ⚠️  REFUSED rename {rule.get('anchor')} {rule.get('id')} "
            f"-> '{rule.get('name')}': {why}")
    renamed = strip_as_food(rows)
    # ⚠️ BEFORE RESOLUTION, so a name that should not be here cannot be moved somewhere
    #    else instead of leaving. See drop_initialism_expansions.
    initialisms = drop_initialism_expansions(rows, by_entry)
    # ⚠️ ALSO BEFORE RESOLUTION, and for the same reason: a name nobody alive writes
    #    should leave rather than be moved to another row. See drop_dead_language_names.
    dead = load_dead_languages()
    dead_gone = drop_dead_language_names(rows, dead)
    symbols = drop_agrovoc_symbols(rows, by_entry)
    moved = resolve_borrowed(rows, superclasses, off_parents)
    # ⚠️ AFTER RESOLUTION, because a strength word that leaves a row should not be counted
    #    against it. See mark_strength.
    rows = annotate(rows, superclasses, off_parents)
    strength_a, strength_b = mark_strength(rows)
    rows = annotate(rows, superclasses, off_parents)

    say(f"\nentries: {len(rows):,}")
    for reason, n in dropped.most_common():
        say(f"  removed by kind: {n:5,d}  {reason}")
    marks = collections.Counter(m for r in rows for m in r["cut_by"])
    for name, n in sorted(marks.items()):
        # ⚠️ 'hand' is Andy's call, not a rule in ingredient_cuts.py, so there is no
        #    recorded count to reconcile against and asking for one is a KeyError.
        recorded = CUTS.CUTS.get(name, {}).get("takes")
        note = f"   (ingredient_cuts.py records {recorded:,})" if recorded else \
               "   (hand removals, reasons in hand_removals.csv)"
        say(f"  marked {name:20s} {n:5,d}{note}")
    n_removals = sum(len(v) for v in removals.values())
    # ⚠️ THE PARENT-IN-THE-CHILD'S-BUCKET CHECK. See seed_keys. A seed that names an entry
    #    another kept row is anchored on is almost always the parent, and it drags that
    #    row's whole name set onto the child. Reported every build, never fatal, because a
    #    deliberate overlap is possible and a person has to say so.
    anchored = {(r["anchor"], str(r["id"])): r["canonical"] for r in rows
                if r["anchor"] and not r["cut_by"]}
    collisions = []
    for row in authored:
        for key in seed_keys(row.get("seed"), by_entry, by_bucket):
            owner = anchored.get((key[0], key[2]))
            if owner and owner.casefold() != row["name"].strip().casefold():
                collisions.append((row["name"].strip(), key[0], key[2], owner))
    for name, src, ident, owner in collisions:
        say(f"  ⚠️  SEED COLLISION: '{name}' seeds {src}:{ident}, which anchors the kept row "
            f"'{owner}'. That is usually the PARENT and it drags its whole name set in.")
    if not collisions:
        say("  seed-collision check: no authored seed names another kept row's anchor")
    say(f"  authored rows read: {len(authored):,}   "
        + (", ".join(r["name"] for r in authored) if authored else "none"))
    for row in authored_rejected:
        say(f"  ⚠️  REJECTED, no reason given: authored row {row.get('id')}. "
            "A row without a reason is not created.")
    say(f"  hand removals read: {n_removals:,} over {len(removals):,} entries")
    for row in rejected:
        say(f"  ⚠️  REJECTED, no reason given: {row.get('anchor')} {row.get('id')} "
            f"{row.get('action')}. A removal without a reason is not applied.")
    if dangling:
        # ⚠️ REPORTED, NEVER FATAL. The commonest cause is a rule already dropping the
        #    entry, which is good news. Silence is what rots, so it prints every build.
        say(f"  ⚠️  {len(dangling)} DANGLING removal(s): the entry is no longer generated, "
            "so the removal did nothing. A rule may have done the job first, or a key "
            "may have shifted.")
        for row in dangling[:15]:
            say(f"        {row['anchor']} {row['id']} {row['action']}  ({row['reason'][:60]})")
    multi = [r for r in rows if len(r["articles"]) > 1]
    say(f"  rows answering for 2+ en.wikipedia articles: {len(multi):5,d}   "
        f"{sum(len(r['articles']) for r in multi):,} articles between them")
    drinks = [r for r in rows if r["drink"]]
    say(f"  admitted by the drinks rule: {len(drinks):5,d}   "
        "(the loose version admits 65 at 60%)")
    families = {strength_split(r["canonical"])[1] for r in rows if r.get("strength_b")}
    say(f"  strength shape: form A {strength_a:5,d} rows   "
        f"form B {sum(1 for r in rows if r.get('strength_b')):,} rows in "
        f"{len(families):,} families")
    marked = sum(len(r["intruders"]) for r in rows)
    say(f"  second primary names marked: {marked:5,d} over "
        f"{sum(1 for r in rows if r['intruders']):,} entries")
    say(f"  renamed off 'as food':   {len(renamed):5,d}   "
        "(19 more are left alone: the bare stem is another row's canonical)")
    say(f"  initialism expansions dropped: {sum(initialisms.values()):5,d} "
        f"from {len(initialisms):,} rows   "
        "(97 names, 54 recipe lines, 0 names left unreachable)")
    say(f"  dead-language names dropped: {sum(dead_gone.values()):5,d} "
        f"from {len(dead_gone):,} rows   "
        f"({len(dead)} codes, Latin deliberately NOT among them, 0 recipe lines)")
    say(f"  AGROVOC symbol names dropped: {sum(symbols.values()):5,d} "
        f"from {len(symbols):,} rows   "
        "(Cu, Al, Be, Ni, UN, EU, TMTD and their expansions; NOT copper or Myanmar)")
    for rule, n in sorted(moved.items()):
        say(f"  names moved to their owner: {n:5,d}  {rule}")
    say(f"  override list size: {len(CUTS.OVERRIDES)}   "
        "⚠️ if this passes a few dozen the anchor rule is drawn in the wrong place")

    # ⚠️ THE ANCHOR-CLAUSE CHECK. Never ship a cut without re-running the overrides.
    index = collections.defaultdict(set)
    for i, row in enumerate(rows):
        index[row["canonical"].casefold()].add(i)
        for text in row["variations"]:
            index[text.casefold()].add(i)
    killed = [t for t in CUTS.OVERRIDES
              if index.get(t.casefold()) and all(rows[i]["cut_by"] for i in index[t.casefold()])]
    if killed:
        raise SystemExit(f"⚠️  A CUT REMOVED AN OVERRIDE: {killed}. Read the anchor clause "
                         "in ingredient_cuts.py.")
    say("  anchor-clause check: every override survives every cut")

    guard_unharvested(out, removals)
    kept, cut, rule2 = write_sheet(rows, out)
    say(f"\nwrote {out}")
    say(f"  ingredients (kept)       {kept:6,d}")
    say(f"  cut, with the rule       {cut:6,d}")
    say(f"  rule 2, decide by hand   {rule2:6,d}")
    return rows


if __name__ == "__main__":
    build()
